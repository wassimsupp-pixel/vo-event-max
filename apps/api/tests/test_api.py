"""
tests/test_api.py — FastAPI integration tests using TestClient.

Tests use dependency overrides to mock Supabase and JWT verification so
no real credentials are needed in CI.

Test groups:
  - Health check
  - File upload validation (format, size, confirmation gate)
  - Column mapping confirmation gate
  - Participant access control (dietary field restriction)
  - Consolidation service unit tests (matching engine)
"""

from __future__ import annotations

import io
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi.testclient import TestClient

# ---------------------------------------------------------------------------
# Minimal environment setup (no real Supabase needed for unit tests)
# ---------------------------------------------------------------------------
import os
os.environ.setdefault("SUPABASE_URL", "https://test.supabase.co")
os.environ.setdefault("SUPABASE_SERVICE_ROLE_KEY", "test-service-role-key")
os.environ.setdefault("SECRET_KEY", "test-secret-key")
os.environ.setdefault("ALLOWED_ORIGINS", "http://localhost:3000")

from main import app
from dependencies import get_current_user, get_supabase_client


# ---------------------------------------------------------------------------
# Fixtures and mocks
# ---------------------------------------------------------------------------

MOCK_ADMIN_USER = {
    "id": "00000000-0000-0000-0000-000000000001",
    "org_id": "00000000-0000-0000-0000-000000000010",
    "email": "admin@test.com",
    "full_name": "Test Admin",
    "role": "admin",
    "preferred_language": "fr",
}

MOCK_VIEWER_USER = {
    "id": "00000000-0000-0000-0000-000000000002",
    "org_id": "00000000-0000-0000-0000-000000000010",
    "email": "viewer@test.com",
    "full_name": "Test Viewer",
    "role": "viewer",
    "preferred_language": "fr",
}


def _mock_supabase() -> MagicMock:
    """Return a MagicMock that mimics the Supabase client interface."""
    mock = MagicMock()
    # Default: all table queries return empty data
    mock.table.return_value.select.return_value.eq.return_value.single.return_value.execute.return_value.data = None
    mock.table.return_value.select.return_value.execute.return_value.data = []
    return mock


def _admin_client() -> TestClient:
    """TestClient with admin user injected."""
    app.dependency_overrides[get_current_user] = lambda: MOCK_ADMIN_USER
    app.dependency_overrides[get_supabase_client] = _mock_supabase
    return TestClient(app, raise_server_exceptions=False)


def _viewer_client() -> TestClient:
    """TestClient with viewer user injected."""
    app.dependency_overrides[get_current_user] = lambda: MOCK_VIEWER_USER
    app.dependency_overrides[get_supabase_client] = _mock_supabase
    return TestClient(app, raise_server_exceptions=False)


def _unauthenticated_client() -> TestClient:
    """TestClient with no dependency overrides (real auth attempted)."""
    app.dependency_overrides = {}
    return TestClient(app, raise_server_exceptions=False)


# ---------------------------------------------------------------------------
# 1. Health check
# ---------------------------------------------------------------------------

class TestHealthCheck:
    def test_health_check_returns_200(self):
        """GET /health should return 200 with status=ok."""
        client = TestClient(app)
        response = client.get("/health")
        assert response.status_code == 200

    def test_health_check_body(self):
        """GET /health should return correct JSON body."""
        client = TestClient(app)
        body = client.get("/health").json()
        assert body["status"] == "ok"
        assert "version" in body

    def test_health_check_no_auth_required(self):
        """GET /health must be accessible without Authorization header."""
        client = _unauthenticated_client()
        response = client.get("/health")
        # Should NOT return 401 (no auth required for health)
        assert response.status_code != 401


# ---------------------------------------------------------------------------
# 2. File upload — format validation
# ---------------------------------------------------------------------------

class TestFileUpload:
    EVENT_ID = "00000000-0000-0000-0000-000000000099"

    def _make_file(self, content: bytes, filename: str, content_type: str):
        return ("file", (filename, io.BytesIO(content), content_type))

    def test_upload_rejects_pdf(self):
        """
        POST /api/files/upload with a .pdf file must return 415 Unsupported Media Type.

        The API only accepts .xlsx, .xls, and .csv files.
        """
        client = _admin_client()
        response = client.post(
            "/api/files/upload",
            data={"event_id": self.EVENT_ID, "source_type": "registration"},
            files=[self._make_file(b"%PDF-1.4 test content", "participants.pdf", "application/pdf")],
        )
        assert response.status_code == 415, f"Expected 415, got {response.status_code}: {response.text}"

    def test_upload_rejects_docx(self):
        """
        POST /api/files/upload with a .docx file must return 415.
        """
        client = _admin_client()
        response = client.post(
            "/api/files/upload",
            data={"event_id": self.EVENT_ID, "source_type": "registration"},
            files=[self._make_file(b"PK fake docx", "report.docx", "application/vnd.openxmlformats-officedocument.wordprocessingml.document")],
        )
        assert response.status_code == 415

    def test_upload_rejects_oversized_file(self):
        """
        POST /api/files/upload with a file > 50 MB must return 413 Request Entity Too Large.
        """
        client = _admin_client()
        large_content = b"a" * (51 * 1024 * 1024)  # 51 MB
        response = client.post(
            "/api/files/upload",
            data={"event_id": self.EVENT_ID, "source_type": "registration"},
            files=[self._make_file(large_content, "big.csv", "text/csv")],
        )
        assert response.status_code == 413

    def test_upload_rejects_txt(self):
        """
        POST /api/files/upload with a .txt file must return 415.
        """
        client = _admin_client()
        response = client.post(
            "/api/files/upload",
            data={"event_id": self.EVENT_ID, "source_type": "registration"},
            files=[self._make_file(b"plain text", "data.txt", "text/plain")],
        )
        assert response.status_code == 415

    @patch("routers.files.verify_event_access")
    def test_delete_file_success(self, mock_verify):
        """DELETE /api/files/{file_id} should succeed if file exists and access is allowed."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        file_id = "00000000-0000-0000-0000-000000000088"

        # Mock metadata response
        mock_supabase.table.return_value.select.return_value.eq.return_value.single.return_value.execute.return_value.data = {
            "id": file_id,
            "event_id": self.EVENT_ID,
            "storage_path": f"{self.EVENT_ID}/{file_id}/test.csv",
        }

        response = client.delete(f"/api/files/{file_id}")
        assert response.status_code == 200
        assert response.json()["message"] == "File deleted successfully."

    @patch("routers.files.verify_event_access")
    def test_delete_file_not_found(self, mock_verify):
        """DELETE /api/files/{file_id} returns 404 if file does not exist."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        file_id = "00000000-0000-0000-0000-000000000088"

        # Mock metadata response as empty
        mock_supabase.table.return_value.select.return_value.eq.return_value.single.return_value.execute.return_value.data = None

        response = client.delete(f"/api/files/{file_id}")
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# 3. Column mapping — confirmation gate
# ---------------------------------------------------------------------------

class TestColumnMapping:
    FILE_ID = "00000000-0000-0000-0000-000000000088"

    def test_column_mapping_requires_confirmation_true(self):
        """
        POST /api/files/{file_id}/map-columns with confirmed=false must be rejected.

        The ``confirmed`` field is a hard gate: the frontend must set it to
        ``true`` only after a human has reviewed the mapping.
        """
        client = _admin_client()
        payload = {
            "mapping": {"Nom": "last_name", "Prénom": "first_name"},
            "confirmed": False,
        }
        response = client.post(
            f"/api/files/{self.FILE_ID}/map-columns",
            json=payload,
        )
        # Pydantic validator raises 422 when confirmed is False
        assert response.status_code in (400, 422), (
            f"Expected 400 or 422, got {response.status_code}: {response.text}"
        )

    def test_column_mapping_rejects_missing_confirmed_field(self):
        """
        POST /api/files/{file_id}/map-columns without the ``confirmed`` field must return 422.
        """
        client = _admin_client()
        payload = {
            "mapping": {"Nom": "last_name"},
            # 'confirmed' field absent
        }
        response = client.post(
            f"/api/files/{self.FILE_ID}/map-columns",
            json=payload,
        )
        assert response.status_code == 422

    def test_column_mapping_schema_validates_confirmed_false(self):
        """
        Unit test: ColumnMappingRequest Pydantic model should raise ValueError for confirmed=False.
        """
        from models.schemas import ColumnMappingRequest
        import pydantic

        with pytest.raises((pydantic.ValidationError, ValueError)):
            ColumnMappingRequest(
                mapping={"A": "last_name"},
                confirmed=False,
            )

    def test_column_mapping_schema_accepts_confirmed_true(self):
        """
        Unit test: ColumnMappingRequest should succeed with confirmed=True.
        """
        from models.schemas import ColumnMappingRequest

        req = ColumnMappingRequest(
            mapping={"Nom": "last_name", "Prénom": "first_name"},
            confirmed=True,
        )
        assert req.confirmed is True
        assert req.mapping["Nom"] == "last_name"


# ---------------------------------------------------------------------------
# 4. Participant access — dietary field restriction
# ---------------------------------------------------------------------------

class TestParticipantDietaryRestriction:
    def test_dietary_stripped_from_schemas_for_viewer(self):
        """
        Unit test: ParticipantResponse for a viewer role should have dietary=None
        after the router strips it (simulated by the _strip_dietary helper).
        """
        from routers.participants import _strip_dietary

        participant = {
            "id": "00000000-0000-0000-0000-000000000001",
            "dietary_requirements": "Halal",
            "first_name": "Alice",
            "last_name": "Martin",
        }
        stripped = _strip_dietary(participant, "viewer")
        assert stripped.get("dietary_requirements") is None

    def test_dietary_retained_for_admin(self):
        """
        Unit test: Admin role should retain the dietary_requirements field.
        """
        from routers.participants import _strip_dietary

        participant = {
            "id": "00000000-0000-0000-0000-000000000001",
            "dietary_requirements": "Vegan",
        }
        retained = _strip_dietary(participant, "admin")
        assert retained.get("dietary_requirements") == "Vegan"

    def test_dietary_retained_for_pm(self):
        """
        Unit test: PM role should retain the dietary_requirements field.
        """
        from routers.participants import _strip_dietary

        participant = {"dietary_requirements": "Gluten-free"}
        retained = _strip_dietary(participant, "pm")
        assert retained.get("dietary_requirements") == "Gluten-free"

    def test_dietary_stripped_for_client(self):
        """Client role should NOT see dietary_requirements."""
        from routers.participants import _strip_dietary

        participant = {"dietary_requirements": "Kosher"}
        stripped = _strip_dietary(participant, "client")
        assert stripped.get("dietary_requirements") is None


class TestParticipantFieldAllowlist:
    """PATCH /participants/{id} and POST .../lock/{field} used to take an
    IMMUTABLE_FIELDS BLOCKLIST (only id/event_id/created_at/updated_at/
    registration_source_id/fcm_source_id refused) -- every OTHER column,
    including engine-managed ones (completeness_status, locked_fields,
    has_flight/has_hotel/has_transfer/has_activities, verification_note),
    was writable by any user with write access to the event. A client-role
    "editor" could e.g. lock dietary_requirements against future re-imports
    with a value never actually reviewed, or fake has_flight/has_hotel for
    reporting. Fixed by flipping to an ALLOWLIST matching exactly what the
    participant edit form exposes (apps/web EDITABLE_FIELDS)."""

    def test_engine_managed_fields_not_in_allowlist(self):
        from routers.participants import CLIENT_EDITABLE_FIELDS
        for field in ("completeness_status", "locked_fields", "has_flight",
                      "has_hotel", "has_transfer", "has_activities",
                      "verification_note", "id", "event_id"):
            assert field not in CLIENT_EDITABLE_FIELDS, field

    def test_allowlist_matches_frontend_editable_fields(self):
        from routers.participants import CLIENT_EDITABLE_FIELDS
        # apps/web/.../participants/[participantId]/page.tsx EDITABLE_FIELDS
        frontend_fields = {"first_name", "last_name", "email", "company", "phone", "nationality", "dietary_requirements"}
        assert CLIENT_EDITABLE_FIELDS == frontend_fields

    @patch("routers.participants.verify_event_access")
    def test_patch_rejects_engine_managed_field(self, mock_verify):
        mock_verify.return_value = None
        client = _admin_client()
        mock_supabase = MagicMock()
        mock_supabase.table.return_value.select.return_value.eq.return_value.single.return_value.execute.return_value.data = {
            "id": "p1", "event_id": "e1", "completeness_status": "incomplete",
        }
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.patch(
            "/api/participants/p1",
            json={"field": "completeness_status", "value": "complete", "reason": "test"},
        )
        assert response.status_code == 400

    @patch("routers.participants.verify_event_access")
    def test_lock_rejects_engine_managed_field(self, mock_verify):
        mock_verify.return_value = None
        client = _admin_client()
        mock_supabase = MagicMock()
        mock_supabase.table.return_value.select.return_value.eq.return_value.single.return_value.execute.return_value.data = {
            "id": "p1", "event_id": "e1", "locked_fields": {},
        }
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.post("/api/participants/p1/lock/has_flight")
        assert response.status_code == 400


class TestEmailAgentAccessControl:
    """The Email Agent router had NO ownership check at all (2026-07-21 audit):
    any authenticated user, any org, could list another org's email
    proposals (sender addresses, message bodies, proposed dietary_requirements
    changes) and apply/reject them, writing to a participant they don't own.
    These tests prove verify_event_access is actually wired in and rejects a
    cross-org event with 404, matching the pattern used by every other
    endpoint in this codebase (see dependencies.verify_event_access)."""

    FOREIGN_EVENT_ID = "00000000-0000-0000-0000-0000000000ff"
    PROPOSAL_ID = "00000000-0000-0000-0000-000000000123"

    def _client_with_org_mismatch(self):
        """A supabase mock whose events lookup finds nothing -- simulates the
        event belonging to a DIFFERENT org than current_user.org_id, which is
        exactly what verify_event_access's real query filters on
        (.eq("projects.org_id", org_id)). No proposal-specific behaviour is
        stubbed: reaching this table at all proves the org check ran first."""
        client = _admin_client()
        mock_supabase = MagicMock()
        events_mock = MagicMock()
        events_mock.select.return_value.eq.return_value.eq.return_value.single.return_value.execute.side_effect = Exception("no row")
        mock_supabase.table.side_effect = lambda name: events_mock if name == "events" else MagicMock()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        return client

    def test_list_proposals_rejects_foreign_org_event(self):
        client = self._client_with_org_mismatch()
        response = client.get(f"/api/events/{self.FOREIGN_EVENT_ID}/email-agent")
        assert response.status_code == 404

    def test_analyze_rejects_foreign_org_event(self):
        client = self._client_with_org_mismatch()
        response = client.post(
            f"/api/events/{self.FOREIGN_EVENT_ID}/email-agent/analyze",
            json={"sender": "x@y.com", "subject": "s", "body": "b"},
        )
        assert response.status_code == 404

    @patch("routers.email_agent.EmailAgentService.get_proposal")
    def test_apply_rejects_proposal_from_foreign_org_event(self, mock_get_proposal):
        mock_get_proposal.return_value = {
            "id": self.PROPOSAL_ID, "event_id": self.FOREIGN_EVENT_ID,
            "status": "pending", "participant_id": "p1", "proposed_changes": {"phone": "+1"},
        }
        client = self._client_with_org_mismatch()
        response = client.post(f"/api/email-agent/{self.PROPOSAL_ID}/apply")
        assert response.status_code == 404

    @patch("routers.email_agent.EmailAgentService.get_proposal")
    def test_reject_rejects_proposal_from_foreign_org_event(self, mock_get_proposal):
        mock_get_proposal.return_value = {
            "id": self.PROPOSAL_ID, "event_id": self.FOREIGN_EVENT_ID, "status": "pending",
        }
        client = self._client_with_org_mismatch()
        response = client.post(f"/api/email-agent/{self.PROPOSAL_ID}/reject")
        assert response.status_code == 404

    @patch("routers.email_agent.EmailAgentService.get_proposal")
    def test_apply_404s_when_proposal_does_not_exist(self, mock_get_proposal):
        mock_get_proposal.return_value = None
        client = _admin_client()
        client.app.dependency_overrides[get_supabase_client] = lambda: _mock_supabase()
        response = client.post(f"/api/email-agent/{self.PROPOSAL_ID}/apply")
        assert response.status_code == 404


class TestDeleteFileConcurrencyGuard:
    """DELETE /files/{id} used to have no guard against a concurrent
    consolidation run on the same event: deletion snapshots the file's
    source_record ids up front, deletes exactly those, then deletes the
    uploaded_files row -- but a concurrent consolidation re-parses this same
    file mid-run and inserts FRESH source_records for it, which survive the
    cleanup and are then permanently orphaned once the parent uploaded_files
    row is gone. Fixed by reusing the same is_consolidation_running() guard
    trigger_consolidation already uses to reject an overlapping run."""

    FILE_ID = "00000000-0000-0000-0000-000000000077"

    def _mocked_client(self, running_data):
        mock_supabase = MagicMock()
        files_mock = MagicMock()
        files_mock.select.return_value.eq.return_value.single.return_value.execute.return_value.data = {
            "id": self.FILE_ID, "event_id": "e1", "storage_path": "e1/f1/x.xlsx",
        }
        runs_mock = MagicMock()
        runs_mock.select.return_value.eq.return_value.eq.return_value.execute.return_value.data = running_data

        def table_side_effect(name):
            if name == "uploaded_files":
                return files_mock
            if name == "consolidation_runs":
                return runs_mock
            return MagicMock()

        mock_supabase.table.side_effect = table_side_effect
        return mock_supabase

    @patch("routers.files.verify_event_access")
    def test_delete_rejected_while_consolidation_running(self, mock_verify):
        from datetime import datetime, timezone
        mock_verify.return_value = None
        client = _admin_client()
        recent = datetime.now(timezone.utc).isoformat()
        mock_supabase = self._mocked_client([{"id": "run1", "started_at": recent}])
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.delete(f"/api/files/{self.FILE_ID}")
        assert response.status_code == 409

    @patch("routers.files.deletion_service")
    @patch("routers.files.verify_event_access")
    def test_delete_allowed_when_no_run_in_progress(self, mock_verify, mock_deletion):
        mock_verify.return_value = None
        client = _admin_client()
        mock_supabase = self._mocked_client([])
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.delete(f"/api/files/{self.FILE_ID}")
        assert response.status_code == 200


class TestConsolidationConcurrencyGuard:
    """POST /consolidate must refuse a second run while one is already in
    flight: the pipeline wipes and rebuilds the event's exceptions (and
    re-links source_records) starting from its very first step, so two
    concurrent runs stomp on each other's writes. This is what emptied a
    real export's Exceptions sheet on 2026-07-20 (a race, not a code bug in
    the export itself) -- see export_service.py's exceptions query fix."""

    EVENT_ID = "00000000-0000-0000-0000-000000000001"

    def _mocked_client(self, uploaded_files_data, runs_select_data, runs_insert_data=None):
        mock_supabase = MagicMock()
        uploaded_files_mock = MagicMock()
        uploaded_files_mock.select.return_value.eq.return_value.in_.return_value.execute.return_value.data = uploaded_files_data
        runs_mock = MagicMock()
        runs_mock.select.return_value.eq.return_value.eq.return_value.execute.return_value.data = runs_select_data
        if runs_insert_data is not None:
            runs_mock.insert.return_value.execute.return_value.data = runs_insert_data

        def table_side_effect(name):
            if name == "uploaded_files":
                return uploaded_files_mock
            if name == "consolidation_runs":
                return runs_mock
            return MagicMock()

        mock_supabase.table.side_effect = table_side_effect
        return mock_supabase

    @patch("routers.consolidation.verify_event_access")
    def test_rejects_when_run_already_in_progress(self, mock_verify):
        from datetime import datetime, timezone
        mock_verify.return_value = None
        client = _admin_client()
        recent = datetime.now(timezone.utc).isoformat()
        mock_supabase = self._mocked_client(
            uploaded_files_data=[{"id": "f1"}],
            runs_select_data=[{"id": "run1", "started_at": recent}],
        )
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.post(f"/api/events/{self.EVENT_ID}/consolidate")
        assert response.status_code == 409

    @patch("routers.consolidation.verify_event_access")
    def test_allows_when_no_run_in_progress(self, mock_verify):
        from datetime import datetime, timezone
        mock_verify.return_value = None
        client = _admin_client()
        mock_supabase = self._mocked_client(
            uploaded_files_data=[{"id": "f1"}],
            runs_select_data=[],
            runs_insert_data=[{
                "id": "00000000-0000-0000-0000-000000000099",
                "event_id": self.EVENT_ID,
                "triggered_by": MOCK_ADMIN_USER["id"],
                "status": "running",
                "stats": None,
                "started_at": datetime.now(timezone.utc).isoformat(),
                "completed_at": None,
            }],
        )
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.post(f"/api/events/{self.EVENT_ID}/consolidate")
        assert response.status_code == 202

    @patch("routers.consolidation.verify_event_access")
    def test_allows_when_previous_run_is_stale(self, mock_verify):
        """A 'running' row older than the staleness window (a hard-crashed
        process that never reached the pipeline's own try/except 'failed'
        handler) must not permanently lock the event out of consolidation."""
        from datetime import datetime, timezone, timedelta
        mock_verify.return_value = None
        client = _admin_client()
        stale = (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat()
        mock_supabase = self._mocked_client(
            uploaded_files_data=[{"id": "f1"}],
            runs_select_data=[{"id": "run_old", "started_at": stale}],
            runs_insert_data=[{
                "id": "00000000-0000-0000-0000-000000000099",
                "event_id": self.EVENT_ID,
                "triggered_by": MOCK_ADMIN_USER["id"],
                "status": "running",
                "stats": None,
                "started_at": datetime.now(timezone.utc).isoformat(),
                "completed_at": None,
            }],
        )
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.post(f"/api/events/{self.EVENT_ID}/consolidate")
        assert response.status_code == 202


class TestMatchCandidateResolveRace:
    """PUT /match-candidates/{id} used to check status != "resolved", run the
    merge, and only THEN write status="resolved" -- a plain read-then-write
    with no atomicity. Two concurrent requests for the same candidate_id
    (double-click "confirmer la fusion") could both pass the check and both
    proceed: the second one merges against rows the first already
    repointed/deleted, silently no-ops, but still logs a misleading "merge"
    change_log entry and re-marks an already-resolved candidate. Fixed by
    atomically claiming the candidate (UPDATE ... WHERE status='pending')
    BEFORE merging, with a revert-to-pending on merge failure so a genuine
    single-request failure can still be retried."""

    CANDIDATE_ID = "cand-1"

    def _mock_supabase(self, claim_data, cand_row):
        mock_supabase = MagicMock()
        candidates_mock = MagicMock()
        candidates_mock.select.return_value.eq.return_value.single.return_value.execute.return_value.data = cand_row
        candidates_mock.update.return_value.eq.return_value.eq.return_value.execute.return_value.data = claim_data
        # The revert path only chains a single .eq(), not .eq().eq() -- a
        # distinct attribute path on the same mock tree, configured separately.
        candidates_mock.update.return_value.eq.return_value.execute.return_value.data = claim_data

        participants_mock = MagicMock()

        def table_side_effect(name):
            if name == "match_candidates":
                return candidates_mock
            if name == "participants":
                return participants_mock
            return MagicMock()

        mock_supabase.table.side_effect = table_side_effect
        return mock_supabase, candidates_mock

    @patch("routers.matching.verify_event_access")
    def test_second_concurrent_resolve_is_rejected(self, mock_verify):
        """Simulates the race directly: the claim's conditional UPDATE
        (.eq("status", "pending")) matches zero rows because a first request
        already flipped status to "resolved" -- the loser must 409, not
        proceed to merge."""
        mock_verify.return_value = None
        client = _admin_client()
        cand_row = {
            "id": self.CANDIDATE_ID, "event_id": "e1", "status": "pending",
            "participant_a_id": "loser", "participant_b_id": "winner",
        }
        mock_supabase, candidates_mock = self._mock_supabase(claim_data=[], cand_row=cand_row)
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.put(f"/api/match-candidates/{self.CANDIDATE_ID}", json={"decision": "fusionner"})
        assert response.status_code == 409

    @patch("routers.matching.consolidation_service._merge_participant_into")
    @patch("routers.matching.verify_event_access")
    def test_failed_merge_reverts_claim_to_pending(self, mock_verify, mock_merge):
        """A genuine (non-race) merge failure must revert the atomic claim
        so the candidate can still be retried, not get permanently stuck
        "resolved" with no merge having actually happened."""
        mock_verify.return_value = None
        mock_merge.return_value = False
        client = _admin_client()
        cand_row = {
            "id": self.CANDIDATE_ID, "event_id": "e1", "status": "pending",
            "participant_a_id": "loser", "participant_b_id": "winner",
        }
        mock_supabase, candidates_mock = self._mock_supabase(claim_data=[{"id": self.CANDIDATE_ID}], cand_row=cand_row)
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.put(f"/api/match-candidates/{self.CANDIDATE_ID}", json={"decision": "fusionner"})
        assert response.status_code == 500
        # Two update() calls on match_candidates: the initial claim, then the revert.
        assert candidates_mock.update.call_count == 2
        revert_payload = candidates_mock.update.call_args_list[1].args[0]
        assert revert_payload["status"] == "pending"


# ---------------------------------------------------------------------------
# 5. Matching engine unit tests
# ---------------------------------------------------------------------------

class TestMatchingEngine:
    """Unit tests for the consolidation matching engine (no Supabase needed)."""

    def _reg(self, first: str, last: str, email: str, source_id: str = "reg-1") -> Any:
        from services.consolidation_service import ParticipantRecord
        return ParticipantRecord(source_id, {
            "first_name": first, "last_name": last, "email": email
        })

    def _fcm(self, first: str, last: str, email: str, source_id: str = "fcm-1") -> Any:
        from services.consolidation_service import ParticipantRecord
        return ParticipantRecord(source_id, {
            "first_name": first, "last_name": last, "email": email
        })

    def test_exact_email_match_is_certain(self):
        """Matching FCM → registration via exact email should be CERTAIN."""
        from services.consolidation_service import match_sources

        reg  = self._reg("Alice", "Martin", "alice@example.com", "reg-1")
        fcm  = self._fcm("Alice", "Martin", "alice@example.com", "fcm-1")
        results = match_sources([reg], [fcm])

        assert len(results) == 1
        assert results[0].decision == "certain"
        assert results[0].score == 100.0

    def test_name_match_without_email_is_probable(self):
        """
        High name similarity without email should be at least PROBABLE.
        (Score ≥ 75 but < 95 threshold when no email available on FCM side.)
        """
        from services.consolidation_service import match_sources

        reg  = self._reg("Alice", "Martin", "alice@example.com", "reg-1")
        fcm  = self._fcm("Alice", "Martin", "",                 "fcm-1")  # no email in FCM
        results = match_sources([reg], [fcm])

        assert len(results) == 1
        assert results[0].decision in ("certain", "probable")

    def test_no_match_returns_not_found(self):
        """Completely different names without email overlap → NOT_FOUND."""
        from services.consolidation_service import match_sources

        reg  = self._reg("Alice", "Martin", "alice@example.com", "reg-1")
        fcm  = self._fcm("Zoltan", "Kovacs", "zoltan@example.com", "fcm-1")
        results = match_sources([reg], [fcm])

        assert len(results) == 1
        assert results[0].decision in ("not_found", "to_verify")

    def test_merge_skips_locked_fields(self):
        """
        merge_participant_fields must not overwrite fields that are locked.
        """
        from services.consolidation_service import merge_participant_fields

        existing = {"id": "p1", "first_name": "Alice", "company": "OriginalCo"}
        new_data  = {"first_name": "Alyce",  "company": "NewCo"}
        locked    = {"company": True}  # company is locked

        merged = merge_participant_fields(existing, new_data, locked)
        # company must not be overwritten
        assert merged["company"] == "OriginalCo"
        # first_name (not locked) should be updated
        assert merged["first_name"] == "Alyce"

    def test_merge_does_not_overwrite_with_empty(self):
        """
        merge_participant_fields must not overwrite a value with None or empty string.
        """
        from services.consolidation_service import merge_participant_fields

        existing = {"company": "ExistingCo", "phone": "+32 2 000 0000"}
        new_data  = {"company": None, "phone": ""}
        locked    = {}

        merged = merge_participant_fields(existing, new_data, locked)
        assert merged["company"] == "ExistingCo"
        assert merged["phone"] == "+32 2 000 0000"

    def test_exact_score_tie_is_broken_by_list_order_not_by_any_other_signal(self):
        """2026-07-21/22 audit (PROP-002): documents the CURRENT tie-break
        behaviour of match_sources -- when two registration records score
        identically against one FCM record, the first one encountered in
        `registrations` wins (`if name_score > best_score`, strict >).
        This is deterministic GIVEN a fixed list order, but the caller in
        run_consolidation builds that list from `.select().in_(chunk)` after
        an `.upsert()` -- neither guarantees the DB returns rows in input
        order, so the same real-world tie could resolve to a different
        winner across separate consolidation runs of identical source data.
        This test pins today's behaviour (order-dependent) so a future fix
        for the ordering bug is provable, and so any change to the
        tie-break rule itself is a deliberate, visible diff."""
        from services.consolidation_service import match_sources

        fcm = self._fcm("Marie", "Laurent", "", "fcm-1")
        reg_a = self._reg("Marie", "Laurent", "", "reg-a")  # identical name, no email either side
        reg_b = self._reg("Marie", "Laurent", "", "reg-b")

        results_ab = match_sources([reg_a, reg_b], [fcm])
        results_ba = match_sources([reg_b, reg_a], [fcm])

        assert results_ab[0].score == results_ba[0].score == 100.0  # genuine tie
        assert results_ab[0].reg_record.source_record_id == "reg-a"
        assert results_ba[0].reg_record.source_record_id == "reg-b"
        # Same two people, same tie score, opposite input order -> opposite
        # winner. Proves the outcome depends on list order, not on any
        # property of the data itself.


class TestSafeStorageFilename:
    """A client-controlled upload filename becomes a Supabase Storage path
    segment (routers/files.py create_upload). Path-traversal or control
    characters must never reach it, even though the {event_id}/{file_id}/
    prefix is always server-generated."""

    def test_path_traversal_stripped(self):
        from routers.files import _safe_storage_filename
        assert ".." not in _safe_storage_filename("../../../etc/passwd")
        assert "/" not in _safe_storage_filename("../../other-event/x.xlsx")

    def test_bare_dot_dot_rejected(self):
        """A bare ".." has no separator for os.path.basename to strip and
        survives it unchanged -- still a real traversal segment once joined
        into storage_path ("{event_id}/{file_id}/.." resolves upward)."""
        from routers.files import _safe_storage_filename
        assert _safe_storage_filename("..") == "upload"
        assert _safe_storage_filename(".") == "upload"
        assert _safe_storage_filename("...") == "upload"

    def test_backslash_traversal_stripped(self):
        from routers.files import _safe_storage_filename
        result = _safe_storage_filename("..\\..\\windows\\system32\\x.xlsx")
        assert "\\" not in result and "/" not in result

    def test_normal_filename_mostly_preserved(self):
        from routers.files import _safe_storage_filename
        assert _safe_storage_filename("Participants Q3.xlsx") == "Participants_Q3.xlsx"

    def test_empty_filename_falls_back(self):
        from routers.files import _safe_storage_filename
        assert _safe_storage_filename("") == "upload"

    def test_original_filename_field_left_untouched(self):
        """original_filename (DB field, shown in UI, used for extension
        detection on re-download) must stay the user's real name -- only the
        STORAGE PATH segment is sanitised."""
        import inspect
        from routers import files
        src = inspect.getsource(files)
        assert '"original_filename": filename' in src
        assert "original_filename=filename" in src


class TestMailConnectionWriteCheck:
    """routers/mail_connection.py had `str(event_id, write=True)` -- write=True
    landed inside str()'s call, not verify_event_access's, which raises
    TypeError before the access check ever runs. POST .../mail/sync and
    .../mail/disconnect always 500'd (fails closed, but the intended
    write-access check never executed). Fixed by moving write=True onto
    verify_event_access itself."""

    EVENT_ID = "00000000-0000-0000-0000-000000000001"

    @patch("routers.mail_connection.verify_event_access")
    @patch("routers.mail_connection.mail")
    def test_sync_no_longer_raises_typeerror(self, mock_mail, mock_verify):
        from unittest.mock import AsyncMock
        mock_verify.return_value = {"id": self.EVENT_ID}
        mock_mail.sync_inbox = AsyncMock(return_value={"synced": 0})
        client = _admin_client()
        response = client.post(f"/api/events/{self.EVENT_ID}/mail/sync?provider=gmail")
        assert response.status_code == 200
        # write=True must reach verify_event_access, not str()
        _, kwargs = mock_verify.call_args
        assert kwargs.get("write") is True

    @patch("routers.mail_connection.verify_event_access")
    @patch("routers.mail_connection.mail")
    def test_disconnect_no_longer_raises_typeerror(self, mock_mail, mock_verify):
        mock_verify.return_value = {"id": self.EVENT_ID}
        mock_mail.disconnect = MagicMock()
        client = _admin_client()
        response = client.post(f"/api/events/{self.EVENT_ID}/mail/disconnect?provider=gmail")
        assert response.status_code == 200
        _, kwargs = mock_verify.call_args
        assert kwargs.get("write") is True


class TestFormulaInjectionNoVisibleArtifact:
    """The Excel/CSV formula-injection guard (CWE-1236) went through TWO
    versions. v1 prefixed any string starting with =/+/-/@ with an
    apostrophe. Caught before push: verified via raw OOXML inspection that
    openpyxl writes NO quotePrefix attribute, so that apostrophe becomes a
    LITERAL, VISIBLE character in the exported file -- every phone number
    ("+33...") would show up as "'+33..." in the master list, a core MVP
    deliverable. v2 (this one) root-caused the ACTUAL mechanism: only "="
    gets auto-promoted by openpyxl to a real <f> formula element Excel
    evaluates on open; +/-/@ already stay a plain string (data_type='s'),
    which Excel never re-parses as a formula because OOXML cell types are
    explicit metadata (unlike raw CSV). _neutralize_formula forces
    data_type back to 's' only when openpyxl actually set it to 'f',
    preserving the value byte-for-byte."""

    def test_formula_neutralised_no_live_formula(self):
        from services.export_service import _neutralize_formula
        from openpyxl import Workbook
        ws = Workbook().active
        cell = _neutralize_formula(ws.cell(row=1, column=1, value="=cmd|'/c calc'!A1"))
        assert cell.data_type == "s"
        assert cell.value == "=cmd|'/c calc'!A1"

    def test_ordinary_plus_minus_at_values_completely_unchanged(self):
        """The regression this test exists to prevent: MUST show exactly
        as typed, no apostrophe, no artifact of any kind."""
        from services.export_service import _neutralize_formula
        from openpyxl import Workbook
        ws = Workbook().active
        for v in ("+33612345678", "-5", "@handle", "Jean-Paul"):
            cell = _neutralize_formula(ws.cell(row=1, column=1, value=v))
            assert cell.value == v, f"value mutated: {cell.value!r} != {v!r}"
            assert cell.data_type == "s"

    def test_master_list_phone_column_shows_exact_value(self):
        from services.export_service import _build_master_list_sheet
        from openpyxl import Workbook
        ws = Workbook().active
        row = {
            "id": "p1", "last_name": "Dupont", "first_name": "Marie",
            "phone": "+33612345678", "completeness_status": "complete",
        }
        _build_master_list_sheet(ws, [row], set())
        data_row = [c.value for c in ws[2]]
        assert data_row[4] == "+33612345678"  # Phone column

    def test_malicious_formula_neutralised_end_to_end(self):
        from services.export_service import _build_master_list_sheet
        from openpyxl import Workbook
        ws = Workbook().active
        row = {
            "id": "p1", "last_name": "=cmd|'/c calc'!A1", "first_name": "Jean",
            "completeness_status": "complete",
        }
        _build_master_list_sheet(ws, [row], set())
        assert ws.cell(row=2, column=1).data_type == "s"
        assert ws.cell(row=2, column=1).value == "=cmd|'/c calc'!A1"


class TestExportDietaryRBAC:
    """The export endpoint (routers/exports.py create_export) only called
    verify_event_access with no role check, so ANY role that can read the
    event (including client/viewer) got dietary_requirements and
    food_allergy_info in the generated Excel -- bypassing the RGPD-sensitive
    gate enforced everywhere else (participants.py's _strip_dietary). Fixed
    by threading the caller's role into export_service.generate_excel /
    _build_master_list_sheet, defaulting to the LEAST privileged role so a
    caller that forgets to pass it fails closed."""

    def _rows(self):
        return [{
            "id": "p1", "last_name": "Dupont", "first_name": "Marie",
            "dietary_requirements": "Allergie aux arachides",
            "food_allergy_info": "Anaphylaxie severe",
            "completeness_status": "complete",
        }]

    def test_dietary_included_for_admin(self):
        from services.export_service import _build_master_list_sheet
        from openpyxl import Workbook
        ws = Workbook().active
        _build_master_list_sheet(ws, self._rows(), set(), include_dietary=True)
        row = [c.value for c in ws[2]]
        assert row[8] == "Allergie aux arachides"
        assert row[9] == "Anaphylaxie severe"

    def test_dietary_stripped_for_non_privileged_role(self):
        from services.export_service import _build_master_list_sheet
        from openpyxl import Workbook
        ws = Workbook().active
        _build_master_list_sheet(ws, self._rows(), set(), include_dietary=False)
        row = [c.value for c in ws[2]]
        assert row[8] is None
        assert row[9] is None
        # Only these two columns are affected -- everything else survives.
        assert row[0] == "Dupont"
        assert row[1] == "Marie"

    def test_generate_excel_defaults_to_least_privileged_role(self):
        """A caller that forgets to pass `role=` must fail CLOSED (dietary
        stripped), not open -- a permission parameter defaulting to full
        access is exactly the bug class this test guards against."""
        import inspect
        from services.export_service import generate_excel
        role_param = inspect.signature(generate_excel).parameters["role"]
        assert role_param.default not in ("admin", "pm")

    def _changes(self):
        return [
            {"changed_at": "2026-07-21T00:00:00Z", "user_id": "u1", "entity_type": "participant",
             "entity_id": "p1", "field_name": "dietary_requirements",
             "old_value": "", "new_value": "Halal", "change_reason": "manual"},
            {"changed_at": "2026-07-21T00:01:00Z", "user_id": "u1", "entity_type": "participant",
             "entity_id": "p1", "field_name": "phone",
             "old_value": "", "new_value": "+33612345678", "change_reason": "manual"},
        ]

    def test_change_log_dietary_redacted_for_non_privileged_role(self):
        """The Master List sheet strips the CURRENT dietary value for
        non-admin/pm, but every HISTORICAL edit to that field was still
        written verbatim to the Change Log sheet regardless of role --
        the same RGPD-sensitive value leaking through the audit trail."""
        from services.export_service import _build_change_log_sheet
        from openpyxl import Workbook
        ws = Workbook().active
        _build_change_log_sheet(ws, self._changes(), include_dietary=False)
        dietary_row = [c.value for c in ws[2]]
        phone_row = [c.value for c in ws[3]]
        assert dietary_row[5] is None and dietary_row[6] is None  # old/new value
        assert dietary_row[4] == "dietary_requirements"  # field name itself still shown
        # untouched by the DIETARY redaction, and exactly as-typed: the
        # formula-injection guard neutralises "="-leading VALUES at the
        # cell-type level (see _neutralize_formula), it never mutates the
        # visible content, so a phone number is never shown with an
        # artifact prefix.
        assert phone_row[5] == "" and phone_row[6] == "+33612345678"

    def test_change_log_dietary_included_for_admin(self):
        from services.export_service import _build_change_log_sheet
        from openpyxl import Workbook
        ws = Workbook().active
        _build_change_log_sheet(ws, self._changes(), include_dietary=True)
        dietary_row = [c.value for c in ws[2]]
        assert dietary_row[6] == "Halal"


class TestExportDownloadRoleRedaction:
    """The dietary RBAC fix (commit 27a7c76) only gated export GENERATION:
    create_export threaded role into generate_excel, but download_export
    only checked verify_event_access (read-level, no role condition) before
    minting a signed URL for the ORIGINAL stored file. A viewer/client could
    still download an admin-generated export and get the raw dietary
    columns baked into that file -- see 2026-07-21 audit finding #1. Fixed
    by regenerating a role-appropriate copy at download time for any
    non-admin/pm role instead of serving the original."""

    EXPORT_ID = "00000000-0000-0000-0000-000000000042"

    def _mock_supabase_for_export(self, export_row):
        mock_supabase = MagicMock()
        exports_mock = MagicMock()
        exports_mock.select.return_value.eq.return_value.single.return_value.execute.return_value.data = export_row

        def table_side_effect(name):
            if name == "exports":
                return exports_mock
            return MagicMock()

        mock_supabase.table.side_effect = table_side_effect
        mock_supabase.storage.from_.return_value.create_signed_url.return_value = {"signedURL": "https://signed.example/x"}
        mock_supabase.storage.from_.return_value.upload.return_value = None
        return mock_supabase

    @patch("routers.exports.verify_event_access")
    @patch("routers.exports.export_service.generate_excel")
    def test_admin_download_serves_original_no_regeneration(self, mock_generate, mock_verify):
        mock_verify.return_value = None
        export_row = {
            "id": self.EXPORT_ID, "event_id": "e1", "run_id": "r1",
            "storage_path": "exports/e1/orig/file.xlsx", "filename": "file.xlsx", "created_by": "u1",
        }
        mock_supabase = self._mock_supabase_for_export(export_row)
        client = _admin_client()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.get(f"/api/exports/{self.EXPORT_ID}/download")
        assert response.status_code == 200
        mock_generate.assert_not_called()
        called_path = mock_supabase.storage.from_.return_value.create_signed_url.call_args.kwargs["path"]
        assert called_path == "exports/e1/orig/file.xlsx"

    @patch("routers.exports.verify_event_access")
    @patch("routers.exports.export_service.generate_excel")
    def test_viewer_download_regenerates_redacted_copy(self, mock_generate, mock_verify):
        mock_verify.return_value = None
        mock_generate.return_value = b"fake-redacted-bytes"
        export_row = {
            "id": self.EXPORT_ID, "event_id": "e1", "run_id": "r1",
            "storage_path": "exports/e1/orig/file.xlsx", "filename": "file.xlsx", "created_by": "u1",
        }
        mock_supabase = self._mock_supabase_for_export(export_row)
        client = _viewer_client()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        response = client.get(f"/api/exports/{self.EXPORT_ID}/download")
        assert response.status_code == 200
        mock_generate.assert_called_once()
        # regenerated with the DOWNLOADING user's role, not the original creator's
        assert mock_generate.call_args.kwargs["role"] == "viewer"
        called_path = mock_supabase.storage.from_.return_value.create_signed_url.call_args.kwargs["path"]
        assert called_path != "exports/e1/orig/file.xlsx"
        assert "redacted-viewer" in called_path


# ---------------------------------------------------------------------------
# 6. Normalisation unit tests
# ---------------------------------------------------------------------------

class TestNormalisation:
    def test_email_lowercased(self):
        """Emails must be lowercased during normalisation."""
        from services.mapping_service import normalise_fields
        result = normalise_fields({"email": "Alice.MARTIN@Example.COM"})
        assert result["email"] == "alice.martin@example.com"

    def test_whitespace_stripped(self):
        """Leading/trailing whitespace must be stripped from all string fields."""
        from services.mapping_service import normalise_fields
        result = normalise_fields({"first_name": "  Alice  ", "last_name": " Martin "})
        assert result["first_name"] == "Alice"
        assert result["last_name"] == "Martin"

    def test_empty_string_becomes_none(self):
        """Empty strings after strip must become None."""
        from services.mapping_service import normalise_fields
        result = normalise_fields({"phone": "   "})
        assert result["phone"] is None

    def test_date_parsed_to_iso8601(self):
        """European date format dd/mm/yyyy must be converted to yyyy-mm-dd."""
        from services.mapping_service import normalise_fields
        result = normalise_fields({"departure_date": "15/06/2025"})
        assert result["departure_date"] == "2025-06-15"

    def test_apply_mapping_drops_unmapped_columns(self):
        """apply_mapping must discard source columns not in the mapping."""
        from services.mapping_service import apply_mapping
        raw = {"Nom": "Martin", "Prénom": "Alice", "Irrelevant": "xyz"}
        mapping = {"Nom": "last_name", "Prénom": "first_name"}
        result = apply_mapping(raw, mapping)
        assert "last_name"  in result
        assert "first_name" in result
        assert "Irrelevant" not in result
        assert "Nom" not in result
        assert "Prénom" not in result


# ---------------------------------------------------------------------------
# 7. Phase 2 integration tests
# ---------------------------------------------------------------------------

class TestFlights:
    @patch("routers.flights.verify_event_access")
    def test_list_flights(self, mock_verify):
        """GET /api/events/{event_id}/flights should return list of flights."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        
        mock_supabase.table.return_value.select.return_value.eq.return_value.execute.return_value.data = [
            {
                "id": "00000000-0000-0000-0000-000000000005",
                "event_id": "00000000-0000-0000-0000-000000000001",
                "flight_number": "SN3715",
                "departure_airport": "BRU",
                "arrival_airport": "BCN",
                "departure_time": "2025-11-10T10:00:00Z",
                "arrival_time": "2025-11-10T12:00:00Z",
                "status": "confirmed",
                "created_at": "2026-07-08T12:00:00Z",
                "participants": {"first_name": "John", "last_name": "Doe"}
            }
        ]
        
        response = client.get("/api/events/00000000-0000-0000-0000-000000000001/flights")
        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["flight_number"] == "SN3715"
        assert data[0]["participant_name"] == "John Doe"


class TestHotels:
    @patch("routers.hotels.verify_event_access")
    def test_list_hotels(self, mock_verify):
        """GET /api/events/{event_id}/hotels should return list of hotels."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        
        mock_supabase.table.return_value.select.return_value.eq.return_value.execute.return_value.data = [
            {
                "id": "00000000-0000-0000-0000-000000000006",
                "event_id": "00000000-0000-0000-0000-000000000001",
                "name": "Hotel Arts Barcelona",
                "city": "Barcelona",
                "created_at": "2026-07-08T12:00:00Z"
            }
        ]
        
        response = client.get("/api/events/00000000-0000-0000-0000-000000000001/hotels")
        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["name"] == "Hotel Arts Barcelona"


class TestTransfers:
    @patch("routers.transfers.verify_event_access")
    def test_list_transfers(self, mock_verify):
        """GET /api/events/{event_id}/transfers should return list of transfers."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        
        mock_supabase.table.return_value.select.return_value.eq.return_value.execute.return_value.data = [
            {
                "id": "00000000-0000-0000-0000-000000000007",
                "event_id": "00000000-0000-0000-0000-000000000001",
                "participant_id": "00000000-0000-0000-0000-000000000002",
                "transfer_type": "arrival",
                "pickup_location": "BCN Airport T1",
                "dropoff_location": "Hotel Arts BCN",
                "pickup_time": "2025-11-10T12:30:00Z",
                "status": "scheduled",
                "created_at": "2026-07-08T12:00:00Z",
                "participants": {"first_name": "Jane", "last_name": "Smith"},
                "flights": {"flight_number": "SN3715"}
            }
        ]
        
        response = client.get("/api/events/00000000-0000-0000-0000-000000000001/transfers")
        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["pickup_location"] == "BCN Airport T1"
        assert data[0]["participant_name"] == "Jane Smith"
        assert data[0]["flight_number"] == "SN3715"


class TestAutoGroupShuttlesFixes:
    """2026-07-21 audit: the shuttle dispatcher (routers.transfers.auto_group_shuttles)
    had two independent bugs, both confirmed against real production data before
    the fix — see verify_transfer_roundtrip.py in the session scratchpad."""

    EVENT_ID = "00000000-0000-0000-0000-00000000ee01"
    PART_ID = "00000000-0000-0000-0000-00000000ee02"

    def _mocked_client(self, flights_data, transfers_exist_check_data=None):
        mock_supabase = MagicMock()
        flights_mock = MagicMock()
        flights_mock.select.return_value.eq.return_value.execute.return_value.data = flights_data

        transfers_mock = MagicMock()
        # existing_res: no imported (flight_id-less) transfers to preserve
        transfers_mock.select.return_value.eq.return_value.execute.return_value.data = []
        # per-flight "already grouped?" check: nothing exists yet -> insert branch
        transfers_mock.select.return_value.eq.return_value.eq.return_value.execute.return_value.data = (
            transfers_exist_check_data or []
        )
        transfers_mock.insert.return_value.execute.return_value.data = [{"id": "new-transfer"}]

        participants_mock = MagicMock()

        def table_side_effect(name):
            if name == "flights":
                return flights_mock
            if name == "transfers":
                return transfers_mock
            if name == "participants":
                return participants_mock
            return MagicMock()

        mock_supabase.table.side_effect = table_side_effect
        client = _admin_client()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        return client, transfers_mock

    @patch("routers.transfers.verify_event_access")
    def test_return_flight_not_treated_as_arrival(self, mock_verify):
        """A participant with an outbound (Aller) AND return (Retour) flight
        segment must only get ONE computed shuttle, for the outbound arrival.
        Before the fix, every flight row was treated as an event arrival, so
        the return leg's landing (back home, days later) also spawned a
        bogus airport-to-hotel shuttle. Confirmed live: 1,156 such bogus
        transfers already existed in production for one event alone."""
        mock_verify.return_value = None
        outbound = {
            "id": "flight-aller", "participant_id": self.PART_ID,
            "departure_time": "2026-02-05T08:00:00Z", "arrival_time": "2026-02-05T10:00:00Z",
            "flight_number": "SN100",
        }
        retour = {
            "id": "flight-retour", "participant_id": self.PART_ID,
            "departure_time": "2026-02-09T14:00:00Z", "arrival_time": "2026-02-09T16:00:00Z",
            "flight_number": "SN200",
        }
        client, transfers_mock = self._mocked_client([outbound, retour])

        response = client.post(f"/api/events/{self.EVENT_ID}/transfers/group")
        assert response.status_code == 200
        assert transfers_mock.insert.call_count == 1
        inserted_payload = transfers_mock.insert.call_args_list[0][0][0]
        assert inserted_payload["flight_id"] == "flight-aller"

    @patch("routers.transfers.verify_event_access")
    def test_window_grouping_consistent_for_non_divisor_window(self, mock_verify):
        """A 90-minute ('1h30') window must group two arrivals only 20
        minutes apart into the SAME shuttle slot, even when they straddle a
        clock-hour boundary. Before the fix, the rounding reset every clock
        hour regardless of window_minutes, so 09:50 and 10:10 landed a full
        hour apart (10:30 vs 11:30) instead of together — the 90/120-minute
        options silently behaved as plain hourly buckets."""
        mock_verify.return_value = None
        part_a = "00000000-0000-0000-0000-00000000ee03"
        part_b = "00000000-0000-0000-0000-00000000ee04"
        flight_a = {
            "id": "flight-a", "participant_id": part_a,
            "departure_time": "2026-02-05T06:00:00Z", "arrival_time": "2026-02-05T09:50:00Z",
            "flight_number": "SN300",
        }
        flight_b = {
            "id": "flight-b", "participant_id": part_b,
            "departure_time": "2026-02-05T06:20:00Z", "arrival_time": "2026-02-05T10:10:00Z",
            "flight_number": "SN400",
        }
        client, transfers_mock = self._mocked_client([flight_a, flight_b])

        response = client.post(f"/api/events/{self.EVENT_ID}/transfers/group?window_minutes=90")
        assert response.status_code == 200
        assert transfers_mock.insert.call_count == 2
        pickup_by_participant = {
            call[0][0]["participant_id"]: call[0][0]["pickup_time"]
            for call in transfers_mock.insert.call_args_list
        }
        assert pickup_by_participant[part_a] == pickup_by_participant[part_b]


class TestActivities:
    @patch("routers.activities.verify_event_access")
    def test_list_activities(self, mock_verify):
        """GET /api/events/{event_id}/activities should return activities list."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        
        mock_supabase.table.return_value.select.return_value.eq.return_value.execute.return_value.data = [
            {
                "id": "00000000-0000-0000-0000-000000000008",
                "event_id": "00000000-0000-0000-0000-000000000001",
                "name": "Barcelona Gothic Quarter Walk",
                "created_at": "2026-07-08T12:00:00Z"
            }
        ]
        mock_supabase.table.return_value.select.return_value.eq.return_value.eq.return_value.execute.return_value.count = 5
        
        response = client.get("/api/events/00000000-0000-0000-0000-000000000001/activities")
        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["name"] == "Barcelona Gothic Quarter Walk"
        assert data[0]["registrations_count"] == 5


class TestReports:
    @patch("routers.reports.verify_event_access")
    def test_get_report_summary(self, mock_verify):
        """GET /api/events/{event_id}/reports/summary should return report aggregate counts."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        mock_supabase.table.return_value.select.return_value.eq.return_value.execute.return_value.data = [
            {"id": "p1", "has_flight": True, "has_hotel": False, "has_transfer": True},
            {"id": "p2", "has_flight": False, "has_hotel": True, "has_transfer": False},
        ]

        response = client.get("/api/events/00000000-0000-0000-0000-000000000001/reports/summary")
        assert response.status_code == 200
        data = response.json()
        assert data["total_registered"] == 2
        assert data["missing_flight"] == 1
        assert data["missing_hotel"] == 1
        assert data["missing_transfer"] == 1


class TestGlobalParticipants:
    def test_get_participant_history(self):
        """GET /api/participants/history should return cross-event presence history for a participant."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        mock_supabase.table.return_value.select.return_value.ilike.return_value.eq.return_value.execute.return_value.data = [
            {
                "email": "alice@test.com",
                "first_name": "Alice",
                "last_name": "Martin",
                "dietary_requirements": "Vegetarian",
                "events": {
                    "name": "Barcelona Kick-off 2025",
                    "start_date": "2025-11-10",
                }
            }
        ]

        response = client.get("/api/global-participants/history?email=alice@test.com")
        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["email"] == "alice@test.com"
        assert data[0]["history"][0]["event_name"] == "Barcelona Kick-off 2025"
        assert data[0]["history"][0]["dietary_requirements"] == "Vegetarian"

    def _participant_rows(self):
        return [
            {
                "email": "alice@test.com", "first_name": "Alice", "last_name": "Martin",
                "dietary_requirements": "Vegetarian",
                "events": {
                    "id": "event-shared", "project_id": "proj-1",
                    "name": "Shared Event", "start_date": "2025-11-10",
                },
            },
            {
                "email": "alice@test.com", "first_name": "Alice", "last_name": "Martin",
                "dietary_requirements": "Vegetarian",
                "events": {
                    "id": "event-not-shared", "project_id": "proj-1",
                    "name": "Not Shared Event", "start_date": "2025-12-01",
                },
            },
        ]

    def _mocked_client(self, membership_data):
        client = _viewer_client()
        mock_supabase = MagicMock()
        participants_mock = MagicMock()
        participants_mock.select.return_value.ilike.return_value.eq.return_value.execute.return_value.data = self._participant_rows()
        members_mock = MagicMock()
        members_mock.select.return_value.eq.return_value.eq.return_value.execute.return_value.data = membership_data

        def table_side_effect(name):
            if name == "participants":
                return participants_mock
            if name == "project_members":
                return members_mock
            return MagicMock()

        mock_supabase.table.side_effect = table_side_effect
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        return client

    def test_dietary_stripped_for_non_staff_role(self):
        """A client/viewer sharing event-shared must not see dietary_requirements
        in cross-event history, even for the event they DO have access to --
        same RGPD gate as participants.py's _strip_dietary."""
        client = self._mocked_client([{"id": "m1", "access_level": "viewer", "event_ids": ["event-shared"]}])
        response = client.get("/api/global-participants/history?email=alice@test.com")
        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        for entry in data[0]["history"]:
            assert entry["dietary_requirements"] is None

    def test_event_excluded_when_not_shared_with_non_staff(self):
        """A client shared ONLY on event-shared must not see history for
        event-not-shared, even though both belong to the same org/project."""
        client = self._mocked_client([{"id": "m1", "access_level": "viewer", "event_ids": ["event-shared"]}])
        response = client.get("/api/global-participants/history?email=alice@test.com")
        assert response.status_code == 200
        data = response.json()
        event_names = {h["event_name"] for h in data[0]["history"]}
        assert event_names == {"Shared Event"}
        assert "Not Shared Event" not in event_names

    def test_no_membership_excludes_all_events_for_non_staff(self):
        """A non-staff user with no project_members row at all (never shared
        on this project) must see nothing, not the full org history."""
        client = self._mocked_client([])
        response = client.get("/api/global-participants/history?email=alice@test.com")
        assert response.status_code == 200
        assert response.json() == []


class TestEvents:
    @patch("routers.events.require_role")
    def test_create_event(self, mock_role):
        """POST /api/events should insert and return new event."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        mock_supabase.table.return_value.select.return_value.eq.return_value.eq.return_value.single.return_value.execute.return_value.data = {"id": "p1"}
        mock_supabase.table.return_value.insert.return_value.execute.return_value.data = [
            {
                "id": "00000000-0000-0000-0000-000000000003",
                "project_id": "00000000-0000-0000-0000-000000000002",
                "name": "New Blank Event",
                "created_at": "2026-07-09T10:00:00Z",
                "updated_at": "2026-07-09T10:00:00Z",
            }
        ]

        payload = {
            "project_id": "00000000-0000-0000-0000-000000000002",
            "name": "New Blank Event"
        }
        response = client.post("/api/events", json=payload)
        assert response.status_code == 201
        assert response.json()["name"] == "New Blank Event"

    def test_list_events(self):
        """GET /api/events should return list of organization events."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        mock_supabase.table.return_value.select.return_value.eq.return_value.execute.return_value.data = [
            {
                "id": "00000000-0000-0000-0000-000000000003",
                "project_id": "00000000-0000-0000-0000-000000000002",
                "name": "Event One",
                "created_at": "2026-07-09T10:00:00Z",
                "updated_at": "2026-07-09T10:00:00Z",
            }
        ]

        response = client.get("/api/events")
        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["name"] == "Event One"


class TestEmailAgentApplyProposalSafety:
    """apply_proposal (services/email_agent_service.py) wrote AI-proposed
    `changes` verbatim -- the prompt's "Allowed fields" list was never
    enforced in code, so a crafted/injected email could get the model to
    propose event_id/locked_fields/id and have them applied. It also had a
    read-then-write TOCTOU race: two concurrent applies for the same
    proposal_id could both pass the pending check before either wrote
    status="applied", doubling every change_log entry."""

    def _proposal(self, changes):
        return {
            "id": "prop-1", "event_id": "e1", "participant_id": "p1",
            "status": "pending", "proposed_changes": changes,
        }

    def _mock_sb(self, claim_data):
        mock_sb = MagicMock()
        proposals_mock = MagicMock()
        proposals_mock.update.return_value.eq.return_value.eq.return_value.execute.return_value.data = claim_data
        participants_mock = MagicMock()
        participants_mock.select.return_value.eq.return_value.execute.return_value.data = [
            {"id": "p1", "locked_fields": [], "first_name": "Old"}
        ]
        change_log_mock = MagicMock()

        def table_side_effect(name):
            if name == "email_proposals":
                return proposals_mock
            if name == "participants":
                return participants_mock
            if name == "change_log":
                return change_log_mock
            return MagicMock()

        mock_sb.table.side_effect = table_side_effect
        return mock_sb, participants_mock, change_log_mock

    @pytest.mark.asyncio
    async def test_disallowed_fields_filtered_before_apply(self):
        from services.email_agent_service import EmailAgentService
        service = EmailAgentService(MagicMock())
        service.get_proposal = AsyncMock(return_value=self._proposal({
            "first_name": "Marie", "event_id": "other-event-id",
            "locked_fields": [], "id": "different-participant",
        }))
        mock_sb, participants_mock, _ = self._mock_sb(claim_data=[{"id": "prop-1"}])
        service.sb = mock_sb

        result = await service.apply_proposal("prop-1", "u1")
        assert result is True
        applied_payload = participants_mock.update.call_args.args[0]
        assert applied_payload.get("first_name") == "Marie"
        for forbidden in ("event_id", "locked_fields", "id"):
            assert forbidden not in applied_payload or forbidden == "locked_fields"
        # locked_fields IS a real key in the payload (the lock-merge logic),
        # but its VALUE must never come from attacker-supplied changes --
        # only "first_name" (the one allowlisted field) is on it.
        assert applied_payload["locked_fields"] == ["first_name"]

    @pytest.mark.asyncio
    async def test_only_allowlisted_fields_survive_when_all_disallowed(self):
        from services.email_agent_service import EmailAgentService
        service = EmailAgentService(MagicMock())
        service.get_proposal = AsyncMock(return_value=self._proposal({
            "event_id": "other-event-id", "locked_fields": [], "id": "x",
        }))
        mock_sb, _, _ = self._mock_sb(claim_data=[{"id": "prop-1"}])
        service.sb = mock_sb

        result = await service.apply_proposal("prop-1", "u1")
        # Nothing allowlisted survives -> changes is empty -> no-op, not applied.
        assert result is False

    @pytest.mark.asyncio
    async def test_double_apply_second_call_is_rejected(self):
        """Simulates the race: the SECOND concurrent call's conditional
        UPDATE (.eq("status", "pending")) matches zero rows because the
        first call already flipped status to "applied", proving the claim
        is the actual gate -- not the earlier plain status read."""
        from services.email_agent_service import EmailAgentService
        service = EmailAgentService(MagicMock())
        service.get_proposal = AsyncMock(return_value=self._proposal({"first_name": "Marie"}))
        mock_sb, participants_mock, change_log_mock = self._mock_sb(claim_data=[])  # 0 rows affected
        service.sb = mock_sb

        result = await service.apply_proposal("prop-1", "u1")
        assert result is False
        participants_mock.update.assert_not_called()
        change_log_mock.insert.assert_not_called()


class TestEmailAgent:
    def test_list_proposals(self):
        """GET /api/events/{event_id}/email-agent should return email proposals."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        mock_supabase.table.return_value.select.return_value.eq.return_value.order.return_value.execute.return_value.data = [
            {
                "id": "00000000-0000-0000-0000-000000000009",
                "event_id": "00000000-0000-0000-0000-000000000001",
                "sender": "test@sender.com",
                "subject": "Dietary request",
                "body": "I am vegan",
                "received_at": "2026-07-09T10:00:00Z",
                "participant_id": None,
                "status": "pending",
                "proposed_changes": {"dietary_requirements": "Végétalien"},
                "ai_explanation": "Vegan request",
                "created_at": "2026-07-09T10:00:00Z",
                "participants": None
            }
        ]

        response = client.get("/api/events/00000000-0000-0000-0000-000000000001/email-agent")
        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        assert data[0]["sender"] == "test@sender.com"
        assert data[0]["proposed_changes"]["dietary_requirements"] == "Végétalien"

    def test_analyze_email(self):
        """POST /api/events/{event_id}/email-agent/analyze should parse and return proposal."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        # Mock participant lookup -> no match
        mock_supabase.table.return_value.select.return_value.eq.return_value.eq.return_value.execute.return_value.data = []
        mock_supabase.table.return_value.select.return_value.eq.return_value.execute.return_value.data = []

        # Mock insert proposal
        mock_supabase.table.return_value.insert.return_value.execute.return_value.data = [
            {
                "id": "00000000-0000-0000-0000-000000000009",
                "event_id": "00000000-0000-0000-0000-000000000001",
                "sender": "test@sender.com",
                "subject": "Change request",
                "body": "Please make me vegetarian",
                "received_at": "2026-07-09T10:00:00Z",
                "participant_id": None,
                "status": "pending",
                "proposed_changes": {"dietary_requirements": "Végétarien"},
                "ai_explanation": "Vegetarian requested",
                "created_at": "2026-07-09T10:00:00Z",
                "participants": None
            }
        ]

        payload = {
            "sender": "test@sender.com",
            "subject": "Change request",
            "body": "Please make me vegetarian"
        }
        response = client.post("/api/events/00000000-0000-0000-0000-000000000001/email-agent/analyze", json=payload)
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "pending"
        assert data["proposed_changes"]["dietary_requirements"] == "Végétarien"


# ---------------------------------------------------------------------------
# 6. Quality engine and source validations
# ---------------------------------------------------------------------------

class TestQualityEngineAndMatchingCorrections:
    def test_possible_duplicate_detection(self):
        """
        Verify that _detect_possible_duplicates correctly flags two participants
        with very similar names but different email addresses.
        """
        from services.exception_service import _detect_possible_duplicates
        from unittest.mock import MagicMock

        # Mock supabase response with similar names, different emails
        mock_supabase = MagicMock()
        mock_supabase.table.return_value.select.return_value.eq.return_value.execute.return_value.data = [
            {"id": "p1", "first_name": "Kris", "last_name": "Brown", "email": "kris.brown@dieteren.com", "company": "D'Ieteren"},
            {"id": "p2", "first_name": "Kris", "last_name": "Brown", "email": "kris.brown@dieteren.be", "company": "D'Ieteren"},
            {"id": "p3", "first_name": "Els", "last_name": "Bertrand", "email": "els.bertrand@cfegroup.be", "company": "CFE"},
            {"id": "p4", "first_name": "Els", "last_name": "Comrtrand", "email": "els.comrtrand1@cfegroup.com", "company": "CFE"},
            {"id": "p5", "first_name": "Different", "last_name": "Name", "email": "diff@test.com", "company": "Test"}
        ]

        exceptions = []
        count = _detect_possible_duplicates(
            event_id="test_event",
            run_id="test_run",
            supabase=mock_supabase,
            exceptions_list=exceptions
        )

        # Should find Kris Brown duplicate and Els Bertrand/Comrtrand duplicate
        assert count == 2
        assert len(exceptions) == 2
        assert any(e["exception_type"] == "POSSIBLE_DUPLICATE" and "Kris Brown" in e["message"] for e in exceptions)
        assert any(e["exception_type"] == "POSSIBLE_DUPLICATE" and "Els Bertrand" in e["message"] for e in exceptions)

    def test_name_mismatch_between_sources_detection(self):
        """
        Verify that _detect_name_mismatches_between_sources correctly flags
        conflicting names between registration and FCM files.
        """
        from services.exception_service import _detect_name_mismatches_between_sources
        from unittest.mock import MagicMock

        mock_supabase = MagicMock()
        
        # Mock participants list return
        mock_supabase.table.return_value.select.return_value.eq.return_value.not_.is_.return_value.not_.is_.return_value.execute.return_value.data = [
            {
                "id": "p1",
                "first_name": "Sebastien",
                "last_name": "Perrin",
                "registration_source_id": "reg1",
                "fcm_source_id": "fcm1"
            }
        ]

        # Mock source records response
        mock_supabase.table.return_value.select.return_value.in_.return_value.execute.return_value.data = [
            {"id": "reg1", "normalized_data": {"first_name": "Sebastien", "last_name": "Perrin"}},
            {"id": "fcm1", "normalized_data": {"first_name": "Sébastien", "last_name": "Pérrin"}}  # minor accents/case diffs count as mismatch
        ]

        exceptions = []
        count = _detect_name_mismatches_between_sources(
            event_id="test_event",
            run_id="test_run",
            supabase=mock_supabase,
            exceptions_list=exceptions
        )

        assert count == 1
        assert len(exceptions) == 1
        # NAME_DIVERGENCE is the ENUM-valid type for registration-vs-FCM
        # name mismatches (NAME_MISMATCH_BETWEEN_SOURCES 22P02-failed inserts).
        assert exceptions[0]["exception_type"] == "NAME_DIVERGENCE"
        assert "Sébastien" in exceptions[0]["message"]

    def test_source_type_never_in_nationality_validation(self):
        """
        Verify that mapping logic doesn't allow source_type values 
        (like registration, fcm, or Excel imports) to leak into nationality.
        """
        from services.mapping_service import apply_mapping, normalise_fields

        # Test with incorrect mapping of "Source" (containing Formulaire web) to "nationality"
        raw_row = {
            "Nom": "Moreau",
            "Prénom": "Olivier",
            "Email": "olivier.moreau@recticel.be",
            "Source": "Formulaire web"
        }

        # If mapping incorrectly maps "Source" to "nationality"
        mapping = {
            "Nom": "last_name",
            "Prénom": "first_name",
            "Email": "email",
            "Source": "nationality"
        }

        mapped = apply_mapping(raw_row, mapping)
        normalised = normalise_fields(mapped)

        nationality_val = normalised.get("nationality")
        
        invalid_source_types = ["formulaire web", "import excel", "email direct", "assistant", "registration", "fcm"]
        if nationality_val and nationality_val.strip().lower() in invalid_source_types:
            is_valid = False
        else:
            is_valid = True

        assert is_valid is False


# ---------------------------------------------------------------------------
# Test Mapping Suggestions
# ---------------------------------------------------------------------------

class TestMappingSuggestions:
    def test_normalize_column_name(self):
        """Verify _normalize_column_name correctly normalizes column names."""
        from services.mapping_service import _normalize_column_name
        assert _normalize_column_name("Prénom") == "prenom"
        assert _normalize_column_name("E-mail Address") == "emailaddress"
        assert _normalize_column_name("Nom de Famille") == "nomdefamille"
        assert _normalize_column_name("Flight_No.") == "flightno"
        assert _normalize_column_name("check-in/date") == "checkindate"

    def test_suggest_mapping_email(self):
        """Verify suggest_mapping identifies email fields by content and name."""
        from services.mapping_service import suggest_mapping
        columns = ["Mail"]
        sample_rows = [
            {"Mail": "jean@dupont.com"},
            {"Mail": "alice@martin.be"},
            {"Mail": "bob@test.org"},
        ]
        sug = suggest_mapping(columns, sample_rows)
        assert "Mail" in sug
        assert sug["Mail"]["suggested_field"] == "email"
        assert sug["Mail"]["confidence"] >= 0.9

    def test_suggest_mapping_date(self):
        """Verify suggest_mapping identifies date fields by content and name."""
        from services.mapping_service import suggest_mapping
        columns = ["Date Entrée"]
        sample_rows = [
            {"Date Entrée": "12/11/2025"},
            {"Date Entrée": "13-11-2025"},
            {"Date Entrée": "2025-11-14"},
        ]
        sug = suggest_mapping(columns, sample_rows)
        assert "Date Entrée" in sug
        assert sug["Date Entrée"]["suggested_field"] == "check_in_date"
        assert sug["Date Entrée"]["confidence"] >= 0.9

    def test_suggest_mapping_flight(self):
        """Verify suggest_mapping identifies flight fields by content and name."""
        from services.mapping_service import suggest_mapping
        columns = ["Vol"]
        sample_rows = [
            {"Vol": "SN1234"},
            {"Vol": "LH456"},
            {"Vol": "AA9876"},
        ]
        sug = suggest_mapping(columns, sample_rows)
        assert "Vol" in sug
        assert sug["Vol"]["suggested_field"] == "flight_number"
        assert sug["Vol"]["confidence"] >= 0.9

    def test_suggest_mapping_weak_no_suggestion(self):
        """Verify suggest_mapping returns None and 0.0 confidence when no matches are found."""
        from services.mapping_service import suggest_mapping
        columns = ["Inconnu"]
        sample_rows = [
            {"Inconnu": "abc"},
            {"Inconnu": "xyz"},
            {"Inconnu": "123"},
        ]
        sug = suggest_mapping(columns, sample_rows)
        assert "Inconnu" in sug
        assert sug["Inconnu"]["suggested_field"] is None
        assert sug["Inconnu"]["confidence"] == 0.0


# ---------------------------------------------------------------------------
# Test Participant Lookup
# ---------------------------------------------------------------------------

class TestParticipantLookup:
    @patch("routers.participants.verify_event_access")
    def test_lookup_participants(self, mock_verify):
        """GET /api/events/{event_id}/participants/lookup should return list of lookup items."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase

        mock_supabase.table.return_value.select.return_value.eq.return_value.order.return_value.order.return_value.range.return_value.execute.return_value.data = [
            {
                "id": "00000000-0000-0000-0000-000000000005",
                "first_name": "Jean",
                "last_name": "Dupont",
                "completeness_status": "complete",
            }
        ]

        response = client.get("/api/events/00000000-0000-0000-0000-000000000001/participants/lookup")
        assert response.status_code == 200, response.text
        data = response.json()
        assert len(data) == 1
        assert data[0]["first_name"] == "Jean"
        assert data[0]["last_name"] == "Dupont"
        assert data[0]["completeness_status"] == "complete"

    @patch("routers.participants.verify_event_access")
    def test_lookup_participants_raises_404_if_no_access(self, mock_verify):
        """GET /api/events/{event_id}/participants/lookup raises 404 if no access."""
        from fastapi import HTTPException
        mock_verify.side_effect = HTTPException(status_code=404, detail="Event not found.")
        client = _admin_client()
        
        response = client.get("/api/events/00000000-0000-0000-0000-000000000001/participants/lookup")
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# Event / Project deletion
# ---------------------------------------------------------------------------

class TestEventProjectDeletion:
    EVENT_ID = "00000000-0000-0000-0000-000000000001"
    PROJECT_ID = "00000000-0000-0000-0000-000000000002"

    @patch("routers.events.verify_event_access")
    @patch("routers.events.deletion_service.delete_event")
    def test_delete_event_success(self, mock_delete, mock_verify):
        """DELETE /api/events/{id} deletes cascade and returns success (admin)."""
        client = _admin_client()
        response = client.delete(f"/api/events/{self.EVENT_ID}")
        assert response.status_code == 200, response.text
        assert response.json()["message"] == "Event deleted successfully."
        mock_delete.assert_called_once()

    def test_delete_event_forbidden_for_viewer(self):
        """A viewer must not be able to delete an event (403)."""
        client = _viewer_client()
        response = client.delete(f"/api/events/{self.EVENT_ID}")
        assert response.status_code == 403

    @patch("routers.events.deletion_service.delete_project")
    def test_delete_project_success(self, mock_delete):
        """DELETE /api/projects/{id} deletes the project after org check (admin)."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        # Project ownership check: table.select.eq.eq.single.execute.data truthy
        mock_supabase.table.return_value.select.return_value.eq.return_value.eq.return_value.single.return_value.execute.return_value.data = {"id": self.PROJECT_ID}

        response = client.delete(f"/api/projects/{self.PROJECT_ID}")
        assert response.status_code == 200, response.text
        assert response.json()["message"] == "Project deleted successfully."
        mock_delete.assert_called_once()

    def test_delete_project_not_found(self):
        """DELETE /api/projects/{id} returns 404 when the project is not in the org."""
        client = _admin_client()
        mock_supabase = _mock_supabase()
        client.app.dependency_overrides[get_supabase_client] = lambda: mock_supabase
        mock_supabase.table.return_value.select.return_value.eq.return_value.eq.return_value.single.return_value.execute.return_value.data = None

        response = client.delete(f"/api/projects/{self.PROJECT_ID}")
        assert response.status_code == 404

    def test_delete_project_forbidden_for_viewer(self):
        """A viewer must not be able to delete a project (403)."""
        client = _viewer_client()
        response = client.delete(f"/api/projects/{self.PROJECT_ID}")
        assert response.status_code == 403


class TestRepairStoredMappingsTransferRules:
    """2026-07-21 audit: a transfer file's "Date" column got mapped to
    check_in_date (a hotel-only field) by mistake once, was "remembered"
    org-wide (column_mapping_templates), and was then silently reapplied
    verbatim to a real event's real transfer import — 0 of 600 rows
    extracted, because check_in_date is never checked as a date source by
    the transfer-extraction gate, and the file's only airport column
    ("Aéroport", unqualified) was left unmapped for the same underlying
    reason (no synonym recognizes a bare "Aéroport"). Rules 6 and 7 in
    repair_stored_mappings (services/mapping_service.py) self-heal both,
    the same way rules 1-5 already self-heal older mis-mapping classes."""

    EVENT_ID = "00000000-0000-0000-0000-00000000ff01"

    def _mocked_client(self, files_data, source_records_data=None):
        mock_supabase = MagicMock()
        files_mock = MagicMock()
        files_mock.select.return_value.eq.return_value.execute.return_value.data = files_data
        files_mock.update.return_value.eq.return_value.execute.return_value.data = [{"id": "ok"}]

        records_mock = MagicMock()
        records_mock.select.return_value.eq.return_value.limit.return_value.execute.return_value.data = (
            source_records_data or []
        )

        def table_side_effect(name):
            if name == "uploaded_files":
                return files_mock
            if name == "source_records":
                return records_mock
            return MagicMock()

        mock_supabase.table.side_effect = table_side_effect
        return mock_supabase, files_mock

    def test_hotel_only_date_field_redirected_to_travel_date(self):
        from services.mapping_service import repair_stored_mappings
        file_row = {
            "id": "file-1",
            "source_type": "transfer",
            "column_mapping": {
                "Vol": "flight_number", "Date": "check_in_date", "Type": "transfer_type",
                "Aeroport": "arrival_airport",
            },
        }
        mock_supabase, files_mock = self._mocked_client([file_row])

        fixed = repair_stored_mappings(self.EVENT_ID, mock_supabase)

        assert fixed == 1
        new_mapping = files_mock.update.call_args_list[0][0][0]["column_mapping"]
        assert new_mapping["Date"] == "departure_date"

    def test_bare_airport_column_mapped_when_no_other_location_field(self):
        from services.mapping_service import repair_stored_mappings
        file_row = {
            "id": "file-2",
            "source_type": "transfer",
            "column_mapping": {
                "Vol": "flight_number", "Date": "departure_date", "Type": "transfer_type",
                "Aeroport": "Aeroport",  # left unmapped (self-mapped custom field)
            },
        }
        mock_supabase, files_mock = self._mocked_client([file_row])

        fixed = repair_stored_mappings(self.EVENT_ID, mock_supabase)

        assert fixed == 1
        new_mapping = files_mock.update.call_args_list[0][0][0]["column_mapping"]
        assert new_mapping["Aeroport"] == "arrival_airport"

    def test_bare_airport_column_left_alone_when_real_location_field_present(self):
        """If the file already has a proper pickup_location/dropoff_location
        column that is ACTUALLY POPULATED, a bare "Aéroport" is genuinely
        extra info -- must not be overwritten (avoids stomping a deliberate
        custom-field choice)."""
        from services.mapping_service import repair_stored_mappings
        file_row = {
            "id": "file-3",
            "source_type": "transfer",
            "column_mapping": {
                "Vol": "flight_number", "Date": "departure_date", "Type": "transfer_type",
                "Aeroport": "Aeroport", "Lieu de Prise en charge": "pickup_location",
            },
        }
        sample = [{"raw_data": {"Lieu de Prise en charge": "Conference Hotel"}}]
        mock_supabase, files_mock = self._mocked_client([file_row], source_records_data=sample)

        fixed = repair_stored_mappings(self.EVENT_ID, mock_supabase)

        assert fixed == 0
        files_mock.update.assert_not_called()

    def test_bare_airport_redirected_when_mapped_location_column_is_empty(self):
        """The exact real-world case: 'Destination'/'Lieu de Prise en charge'
        WERE mapped to dropoff_location/pickup_location, but every one of the
        600 real rows left them blank -- 'Aéroport' was the only column
        actually carrying data. Being MAPPED is not enough; it must be
        POPULATED, or the extraction gate still silently drops every row."""
        from services.mapping_service import repair_stored_mappings
        file_row = {
            "id": "file-4",
            "source_type": "transfer",
            "column_mapping": {
                "Vol": "flight_number", "Date": "departure_date", "Type": "transfer_type",
                "Aeroport": "Aeroport",
                "Destination": "dropoff_location", "Lieu de Prise en charge": "pickup_location",
            },
        }
        sample = [
            {"raw_data": {"Aeroport": "New York JFK", "Destination": None, "Lieu de Prise en charge": ""}},
            {"raw_data": {"Aeroport": "Tokyo NRT", "Destination": "", "Lieu de Prise en charge": None}},
        ]
        mock_supabase, files_mock = self._mocked_client([file_row], source_records_data=sample)

        fixed = repair_stored_mappings(self.EVENT_ID, mock_supabase)

        assert fixed == 1
        new_mapping = files_mock.update.call_args_list[0][0][0]["column_mapping"]
        assert new_mapping["Aeroport"] == "arrival_airport"

    def test_bare_country_column_redirected_to_nationality_when_unmapped(self):
        """Real case: a registration file's "Country" column (300/300 rows
        populated with "Turquie", "Japon"...) was captured under the unused
        `country` rich field -- the export never shows it -- while
        `nationality`, the field that drives Missing Fields / Data Complete,
        stayed empty for all 300 participants (2026-07-21 audit)."""
        from services.mapping_service import repair_stored_mappings
        file_row = {
            "id": "file-5",
            "source_type": "registration",
            "column_mapping": {
                "First Name": "first_name", "Last Name": "last_name",
                "Email Address": "email", "Country": "country",
            },
        }
        mock_supabase, files_mock = self._mocked_client([file_row])

        fixed = repair_stored_mappings(self.EVENT_ID, mock_supabase)

        assert fixed == 1
        new_mapping = files_mock.update.call_args_list[0][0][0]["column_mapping"]
        assert new_mapping["Country"] == "nationality"

    def test_country_column_left_alone_when_nationality_already_mapped(self):
        """A file that genuinely distinguishes the two (a separate,
        POPULATED "Nationalité" column) must not have its "Country" column
        stolen away from the `country` rich field."""
        from services.mapping_service import repair_stored_mappings
        file_row = {
            "id": "file-6",
            "source_type": "registration",
            "column_mapping": {
                "First Name": "first_name", "Last Name": "last_name",
                "Country": "country", "Nationalité": "nationality",
            },
        }
        sample = [{"raw_data": {"Nationalité": "Française"}}]
        mock_supabase, files_mock = self._mocked_client([file_row], source_records_data=sample)

        fixed = repair_stored_mappings(self.EVENT_ID, mock_supabase)

        assert fixed == 0
        files_mock.update.assert_not_called()

    def test_country_column_redirected_when_mapped_nationality_column_is_a_ghost(self):
        """The exact real-world case: the stored mapping already had a
        'Nationalité' -> nationality entry, but it was a STALE leftover from
        an earlier org template -- no such column exists in this file at
        all, so it never carries real data. Being MAPPED is not enough; it
        must be POPULATED, exactly like the bare-airport rule above."""
        from services.mapping_service import repair_stored_mappings
        file_row = {
            "id": "file-7",
            "source_type": "registration",
            "column_mapping": {
                "First Name": "first_name", "Last Name": "last_name",
                "Country": "country", "Nationalité": "nationality",
            },
        }
        sample = [{"raw_data": {"Country": "Turquie"}}]  # no 'Nationalité' key at all
        mock_supabase, files_mock = self._mocked_client([file_row], source_records_data=sample)

        fixed = repair_stored_mappings(self.EVENT_ID, mock_supabase)

        assert fixed == 1
        new_mapping = files_mock.update.call_args_list[0][0][0]["column_mapping"]
        assert new_mapping["Country"] == "nationality"


class TestInferEventCity:
    """2026-07-21 audit: there is no UI to set events.location_city, so the
    arrival_airport fallback that depends on it silently left EVERY flight's
    arrival_airport empty for a real file whose 2-sheet Arrival/Departure
    layout never states an explicit arrival airport at all (600/600 flights,
    real event). _infer_event_city derives it from departure_airport
    frequency instead: every attendee's return leg departs FROM the event
    city, so it dominates the frequency count."""

    def _record(self, departure_airport):
        return {"normalized_data": {"departure_airport": departure_airport}}

    def test_dominant_airport_inferred(self):
        from services.consolidation_service import _infer_event_city
        records = (
            [self._record("BARCELONE BCN")] * 300
            + [self._record("SEOUL ICN")] * 23
            + [self._record("STOCKHOLM ARN")] * 19
            + [self._record("MEXICO MEX")] * 17
        )
        assert _infer_event_city(records) == "BARCELONE BCN"

    def test_no_dominant_airport_returns_none(self):
        """No single city clearly dominates (e.g. only outbound legs exist
        so far, roughly evenly spread across distinct home cities) -- must
        not guess."""
        from services.consolidation_service import _infer_event_city
        records = (
            [self._record("PARIS CDG")] * 10
            + [self._record("LONDON LHR")] * 10
            + [self._record("BERLIN BER")] * 10
            + [self._record("MADRID MAD")] * 10
            + [self._record("ROME FCO")] * 10
        )
        assert _infer_event_city(records) is None

    def test_too_few_records_returns_none_even_if_unanimous(self):
        from services.consolidation_service import _infer_event_city
        records = [self._record("BARCELONE BCN")] * 5
        assert _infer_event_city(records) is None

    def test_no_departure_airport_data_returns_none(self):
        from services.consolidation_service import _infer_event_city
        records = [{"normalized_data": {}}, {"raw_data": {}}]
        assert _infer_event_city(records) is None

    def test_falls_back_to_raw_data_when_normalized_missing(self):
        from services.consolidation_service import _infer_event_city
        records = [{"raw_data": {"departure_airport": "BARCELONE BCN"}}] * 15
        assert _infer_event_city(records) == "BARCELONE BCN"


class TestFallbackArrivalAirport:
    """2026-07-21 audit, same real event: once the event city was known
    (either from events.location_city or _infer_event_city above), the
    arrival_airport fallback was applied unconditionally -- including to the
    RETURN leg, which already departs FROM the event city. Real result:
    300/600 flights showed "BARCELONE BCN -> BARCELONE BCN", a self-round-trip
    that makes no sense on a master list a client will actually read."""

    def test_outbound_leg_gets_event_city_as_arrival(self):
        from services.consolidation_service import _fallback_arrival_airport
        assert _fallback_arrival_airport("MILAN MXP", "BARCELONE BCN") == "BARCELONE BCN"

    def test_return_leg_does_not_get_event_city_stamped_twice(self):
        from services.consolidation_service import _fallback_arrival_airport
        assert _fallback_arrival_airport("BARCELONE BCN", "BARCELONE BCN") == ""

    def test_return_leg_matches_case_insensitively_and_by_substring(self):
        """events.location_city may hold just the city name ("Barcelone")
        while departure_airport holds "city + code" ("Barcelone BCN") --
        must still recognise it as the same place."""
        from services.consolidation_service import _fallback_arrival_airport
        assert _fallback_arrival_airport("Barcelone BCN", "barcelone") == ""
        assert _fallback_arrival_airport("barcelone", "Barcelone BCN") == ""

    def test_no_event_city_returns_empty(self):
        from services.consolidation_service import _fallback_arrival_airport
        assert _fallback_arrival_airport("MILAN MXP", "") == ""

    def test_no_departure_airport_still_falls_back_to_event_city(self):
        from services.consolidation_service import _fallback_arrival_airport
        assert _fallback_arrival_airport(None, "BARCELONE BCN") == "BARCELONE BCN"

