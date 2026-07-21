"""
services/export_service.py — Excel workbook generator for consolidation exports.

Generates a multi-sheet ``.xlsx`` workbook using openpyxl with:
  Sheet 1 — Master List     : all participants, colour-coded by completeness
  Sheet 2 — Exceptions      : all unresolved exceptions from the run
  Sheet 3 — Summary         : run statistics and metadata
  Sheet 4 — Change Log      : last 500 change entries for the event
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from supabase import Client

from services import master_list_service

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Brand colours
# ---------------------------------------------------------------------------
BRAND_PURPLE   = "806CAF"   # Header background
HEADER_FONT    = "FFFFFF"   # Header text colour
ROW_COMPLETE   = "E3F5EA"   # Green  — participant is complete
ROW_INCOMPLETE = "FEECDF"   # Orange — participant is incomplete
ROW_CONFLICT   = "FBE2E1"   # Red    — participant has data conflicts

SEV_CRITICAL   = "FBE2E1"
SEV_WARNING    = "FEECDF"
SEV_INFO       = "EAF4FB"


def _header_fill() -> PatternFill:
    return PatternFill(start_color=BRAND_PURPLE, end_color=BRAND_PURPLE, fill_type="solid")


def _header_font() -> Font:
    return Font(bold=True, color=HEADER_FONT)


def _row_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Adjust column widths based on cell content."""
    for col in ws.columns:
        max_len = max_width
        col_letter = get_column_letter(col[0].column)
        col_max = min_width
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                col_max = max(col_max, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(col_max + 2, max_width)


def _write_header_row(ws, headers: list[str]) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = _header_fill()
        cell.font   = _header_font()
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


# ---------------------------------------------------------------------------
# Sheet 1 — Master List
# ---------------------------------------------------------------------------

def _build_master_list_sheet(ws, rows: list[dict], conflict_ids: set[str], custom_fields: list[str] | None = None) -> None:
    """Populate the detailed Master List sheet (feedback §6): identity + real
    flight / hotel / transfer / activity / dietary details, not just booleans.
    User-defined custom mapping fields are appended as extra columns."""
    ws.title = "Master List"
    ws.freeze_panes = "A2"
    custom_fields = custom_fields or []

    # Bloc headers (MVP feedback §11): the existing multi-line summary columns
    # (Flights/Transfers/Activities) are kept EXACTLY as they were — nothing
    # that already worked changes — these are ADDED alongside them so a 3+-leg
    # itinerary or a 3rd activity is never lost, only additionally summarised
    # in the two-column form the spec asks for.
    headers = [
        "Last Name", "First Name", "Email", "Company", "Phone", "Nationality",
        "Region", "Category", "Dietary Requirements", "Food / Allergy",
        "Flights (airline · route · times)",
        "Outbound Airline", "Outbound Flight #", "Outbound From", "Outbound To",
        "Outbound Departure", "Outbound Arrival", "Outbound Status",
        "Return Airline", "Return Flight #", "Return From", "Return To",
        "Return Departure", "Return Arrival", "Return Status",
        "Hotel", "Check-in", "Check-out", "Nights", "Room",
        "Transfers",
        "Arrival Transfer Type", "Arrival Transfer Route", "Arrival Transfer Time",
        "Arrival Transfer Vehicle", "Arrival Transfer Status",
        "Return Transfer Type", "Return Transfer Route", "Return Transfer Time",
        "Return Transfer Vehicle", "Return Transfer Status",
        "Activities (name · day/time)",
        "Activity 1", "Activity 1 When", "Activity 1 Location", "Activity 1 Status",
        "Activity 2", "Activity 2 When", "Activity 2 Location", "Activity 2 Status",
        "Status",
        "Confirmation Prepared", "Confirmation Sent", "Sent Date", "Needs Update", "Last Communication",
        "Data Complete", "Missing Fields", "Open Exceptions", "Priority",
        "Sources Used", "Last Updated", "Action Needed",
    ] + list(custom_fields)
    _write_header_row(ws, headers)

    def _yn(v: Any) -> str:
        return "Oui" if v else "Non"

    for row_idx, p in enumerate(rows, start=2):
        custom = p.get("custom") or {}
        values = [
            p.get("last_name"),
            p.get("first_name"),
            p.get("email"),
            p.get("company"),
            p.get("phone"),
            p.get("nationality"),
            p.get("region"),
            p.get("attendee_category"),
            p.get("dietary_requirements"),
            p.get("food_allergy_info"),
            p.get("flight_summary"),
            p.get("outbound_airline"), p.get("outbound_flight_number"),
            p.get("outbound_departure_airport"), p.get("outbound_arrival_airport"),
            p.get("outbound_departure_time"), p.get("outbound_arrival_time"), p.get("outbound_status"),
            p.get("return_airline"), p.get("return_flight_number"),
            p.get("return_departure_airport"), p.get("return_arrival_airport"),
            p.get("return_departure_time"), p.get("return_arrival_time"), p.get("return_status"),
            p.get("hotel_name"),
            p.get("hotel_checkin"),
            p.get("hotel_checkout"),
            p.get("hotel_nights_count") or "",
            p.get("hotel_room_type"),
            p.get("transfer_summary"),
            p.get("transfer_arrival_type"), p.get("transfer_arrival_route"),
            p.get("transfer_arrival_time"), p.get("transfer_arrival_vehicle"), p.get("transfer_arrival_status"),
            p.get("transfer_return_type"), p.get("transfer_return_route"),
            p.get("transfer_return_time"), p.get("transfer_return_vehicle"), p.get("transfer_return_status"),
            p.get("activities_summary"),
            p.get("activity_1_name"), p.get("activity_1_when"),
            p.get("activity_1_location"), p.get("activity_1_status"),
            p.get("activity_2_name"), p.get("activity_2_when"),
            p.get("activity_2_location"), p.get("activity_2_status"),
            p.get("completeness_status", "incomplete"),
            _yn(p.get("comm_confirmation_prepared")), _yn(p.get("comm_confirmation_sent")),
            p.get("comm_sent_date"), _yn(p.get("comm_needs_update")), p.get("comm_last_activity"),
            _yn(p.get("dq_complete")), p.get("dq_missing_fields"),
            p.get("dq_open_exceptions") or 0, p.get("dq_priority"),
            p.get("dq_sources_used"), p.get("dq_last_updated"), p.get("dq_action_needed"),
        ] + [custom.get(cf) for cf in custom_fields]

        # Determine row colour
        p_id = str(p.get("id", ""))
        raw_status = p.get("completeness_status", "incomplete")
        if p_id in conflict_ids:
            fill_hex = ROW_CONFLICT
        elif raw_status == "complete":
            fill_hex = ROW_COMPLETE
        else:
            fill_hex = ROW_INCOMPLETE

        fill = _row_fill(fill_hex)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = fill
            cell.border = _thin_border()
            # Multi-line detail cells wrap; the rest are top-aligned for consistency
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _auto_width(ws, max_width=60)


# ---------------------------------------------------------------------------
# Sheet 2 — Exceptions
# ---------------------------------------------------------------------------

def _build_exceptions_sheet(ws, exceptions: list[dict]) -> None:
    """Populate the Exceptions sheet with unresolved exceptions."""
    ws.title = "Exceptions"
    ws.freeze_panes = "A2"

    headers = ["Type", "Severity", "Participant Name", "Message", "Context"]
    _write_header_row(ws, headers)

    # Severity → colour map
    sev_color = {"critical": SEV_CRITICAL, "warning": SEV_WARNING, "info": SEV_INFO}

    for row_idx, exc in enumerate(exceptions, start=2):
        severity = exc.get("severity", "warning")
        context  = exc.get("context_data")
        ctx_str  = str(context) if context else ""

        # Try to build participant name from context
        p_name = (context or {}).get("participant_name", "—") if context else "—"

        values = [
            exc.get("exception_type"),
            severity,
            p_name,
            exc.get("message"),
            ctx_str,
        ]
        fill = _row_fill(sev_color.get(severity, SEV_WARNING))
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _auto_width(ws, max_width=80)


# ---------------------------------------------------------------------------
# Sheet 3 — Summary
# ---------------------------------------------------------------------------

def _build_summary_sheet(ws, run: dict, user_id: str) -> None:
    """Populate the Summary sheet with run statistics."""
    ws.title = "Summary"

    stats: dict = run.get("stats") or {}
    rows = [
        ("Run ID",               run.get("id")),
        ("Event ID",             run.get("event_id")),
        ("Triggered By",         user_id),
        ("Status",               run.get("status")),
        ("Started At",           run.get("started_at")),
        ("Completed At",         run.get("completed_at")),
        ("Generated At",         datetime.now(timezone.utc).isoformat()),
        ("", ""),  # spacer
        ("— Statistics —",       ""),
        ("Total Source Records", stats.get("total_source_records", 0)),
        ("Matched (Certain)",    stats.get("matched_certain", 0)),
        ("Matched (Probable)",   stats.get("matched_probable", 0)),
        ("To Verify",            stats.get("to_verify", 0)),
        ("Not Found",            stats.get("not_found", 0)),
        ("Participants Created", stats.get("participants_created", 0)),
        ("Participants Updated", stats.get("participants_updated", 0)),
        ("Exceptions Count",     stats.get("exceptions_count", 0)),
    ]

    # Header row
    ws.cell(row=1, column=1, value="Key").font   = _header_font()
    ws.cell(row=1, column=1).fill = _header_fill()
    ws.cell(row=1, column=2, value="Value").font = _header_font()
    ws.cell(row=1, column=2).fill = _header_fill()

    for row_idx, (key, value) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=key).font  = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=str(value) if value is not None else "")

    _auto_width(ws, min_width=20)


# ---------------------------------------------------------------------------
# Sheet 4 — Change Log
# ---------------------------------------------------------------------------

def _build_change_log_sheet(ws, changes: list[dict]) -> None:
    """Populate the Change Log sheet (last 500 entries)."""
    ws.title = "Change Log"
    ws.freeze_panes = "A2"

    headers = ["Changed At", "User ID", "Entity Type", "Entity ID", "Field", "Old Value", "New Value", "Reason"]
    _write_header_row(ws, headers)

    for row_idx, ch in enumerate(changes, start=2):
        values = [
            ch.get("changed_at"),
            ch.get("user_id"),
            ch.get("entity_type"),
            str(ch.get("entity_id", "")),
            ch.get("field_name"),
            ch.get("old_value"),
            ch.get("new_value"),
            ch.get("change_reason"),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet — Data-quality analysis & advice (feedback §15)
# ---------------------------------------------------------------------------

def _build_quality_sheet(ws, analysis: dict) -> None:
    """Data-quality dimensions, missing-data counts and distributions."""
    ws.title = "Analyse Qualité"

    ws.cell(row=1, column=1, value="Indicateur").font = _header_font()
    ws.cell(row=1, column=1).fill = _header_fill()
    ws.cell(row=1, column=2, value="Valeur").font = _header_font()
    ws.cell(row=1, column=2).fill = _header_fill()

    dims = analysis.get("dimensions", {}) or {}
    missing = analysis.get("missing", {}) or {}
    dim_labels = {
        "identite": "Identité (email présent)",
        "contact": "Contact (email ou téléphone)",
        "voyage": "Voyage (vol renseigné)",
        "hebergement": "Hébergement",
        "regime": "Régime alimentaire",
        "passeport": "Passeport",
    }
    rows: list[tuple[str, Any]] = [
        ("Participants", analysis.get("total", 0)),
        ("Score qualité global", f"{analysis.get('quality_score', 0)}/100"),
        ("", ""),
        ("— Dimensions (%) —", ""),
    ]
    rows += [(dim_labels.get(k, k), f"{v}%") for k, v in dims.items()]
    rows += [
        ("", ""),
        ("— Données manquantes —", ""),
        ("Sans email", missing.get("email", 0)),
        ("Sans téléphone", missing.get("phone", 0)),
        ("Sans régime alimentaire", missing.get("dietary_requirements", 0)),
        ("Sans n° passeport", missing.get("passport_number", 0)),
        ("Sans fonction/poste", missing.get("job_title", 0)),
        ("Sans région", missing.get("region", 0)),
        ("", ""),
        ("— Services manquants —", ""),
        ("Sans vol", analysis.get("without_flight", 0)),
        ("Sans hébergement", analysis.get("without_hotel", 0)),
        ("Sans transfert", analysis.get("without_transfer", 0)),
        ("Conflits de données non résolus", analysis.get("conflicts", 0)),
        ("Passeports expirés", analysis.get("passport_expired", 0)),
        ("Passeports expirant (< 6 mois)", analysis.get("passport_expiring", 0)),
    ]

    for i, (k, v) in enumerate(rows, start=2):
        c1 = ws.cell(row=i, column=1, value=k)
        if str(k).startswith("—") or k in ("Participants", "Score qualité global"):
            c1.font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(v) if v != "" else "")

    # Distributions on the right
    def _write_distribution(start_col: int, title: str, dist: dict) -> None:
        ws.cell(row=1, column=start_col, value=title).font = _header_font()
        ws.cell(row=1, column=start_col).fill = _header_fill()
        ws.cell(row=1, column=start_col + 1, value="Nombre").font = _header_font()
        ws.cell(row=1, column=start_col + 1).fill = _header_fill()
        for i, (k, v) in enumerate(list(dist.items())[:25], start=2):
            ws.cell(row=i, column=start_col, value=str(k))
            ws.cell(row=i, column=start_col + 1, value=v)

    _write_distribution(4, "Par région", analysis.get("by_region", {}) or {})
    _write_distribution(7, "Par catégorie", analysis.get("by_category", {}) or {})

    _auto_width(ws, min_width=14, max_width=40)


def _build_advice_sheet(ws, analysis: dict) -> None:
    """Prioritised, data-driven recommendations + the AI narrative (feedback §15)."""
    ws.title = "Conseils"

    # AI narrative on top (if any)
    ai = analysis.get("ai_summary")
    start = 1
    if ai:
        ws.cell(row=1, column=1, value="Synthèse (IA)").font = _header_font()
        ws.cell(row=1, column=1).fill = _header_fill()
        cell = ws.cell(row=2, column=1, value=str(ai))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        ws.row_dimensions[2].height = 90
        start = 4

    headers = ["Priorité", "Recommandation", "Nombre concerné"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=col_idx, value=header)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center")

    sev_label = {"critical": "🔴 Critique", "warning": "🟠 Important", "info": "🔵 Info"}
    sev_color = {"critical": SEV_CRITICAL, "warning": SEV_WARNING, "info": SEV_INFO}
    recs = analysis.get("recommendations", []) or []
    # critical → warning → info
    order = {"critical": 0, "warning": 1, "info": 2}
    recs = sorted(recs, key=lambda r: order.get(r.get("severity"), 3))

    for i, r in enumerate(recs, start=start + 1):
        sev = r.get("severity", "info")
        fill = _row_fill(sev_color.get(sev, SEV_INFO))
        vals = [sev_label.get(sev, sev), r.get("text"), r.get("count")]
        for col_idx, value in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if not recs:
        ws.cell(row=start + 1, column=1, value="Aucun point d'attention détecté — données complètes.")

    _auto_width(ws, max_width=90)


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

async def generate_excel(
    event_id: str,
    run_id: str,
    user_id: str,
    supabase: Client,
) -> bytes:
    """
    Generate a multi-sheet Excel workbook for a completed consolidation run.

    Parameters
    ----------
    event_id:  UUID of the event.
    run_id:    UUID of the consolidation_run.
    user_id:   UUID of the user requesting the export.
    supabase:  Supabase client.

    Returns
    -------
    Raw ``.xlsx`` bytes, ready to be uploaded to Supabase Storage.
    """
    # --- Load all data ---

    # Run metadata
    run_resp = (
        supabase.table("consolidation_runs")
        .select("*")
        .eq("id", run_id)
        .single()
        .execute()
    )
    run: dict = run_resp.data or {}

    # Enriched master rows (identity + real flight/hotel/transfer/activity detail).
    # This export runs server-side as an admin operation; the router enforces who
    # can trigger exports, so dietary/allergy fields are included here.
    master_rows: list[dict] = master_list_service.build_master_rows(supabase, event_id)
    master_rows.sort(key=lambda r: str(r.get("last_name") or "").lower())
    custom_fields = sorted({k for r in master_rows for k in (r.get("custom") or {}).keys()})

    # Data-quality analysis (score, dimensions, distributions, recommendations, AI)
    try:
        analysis: dict = master_list_service.build_analysis(supabase, event_id)
    except Exception as exc:
        logger.warning("Analysis for export failed on event %s: %s", event_id, exc)
        analysis = {}

    # Identify participants with DATA_CONFLICT exceptions
    conflict_resp = (
        supabase.table("exceptions")
        .select("participant_id")
        .eq("event_id", event_id)
        .eq("exception_type", "DATA_CONFLICT")
        .eq("resolved", False)
        .execute()
    )
    conflict_ids: set[str] = {
        str(r["participant_id"])
        for r in (conflict_resp.data or [])
        if r.get("participant_id")
    }

    # Unresolved exceptions for this event. Scoped by event_id (not run_id):
    # the exceptions table is wiped and rebuilt at the START of every
    # consolidation run (see consolidation_service step "1b"), long before
    # detection re-populates it near the end. A run_id-scoped query can race
    # a concurrent/subsequent run and see zero rows even though hundreds
    # exist — event_id + resolved matches the pattern already used by the
    # live "Exceptions et alertes" page, which doesn't have this problem.
    exc_resp = (
        supabase.table("exceptions")
        .select("*")
        .eq("event_id", event_id)
        .eq("resolved", False)
        .order("severity")
        .execute()
    )
    exceptions: list[dict] = exc_resp.data or []

    # Change log (last 500 entries for this event)
    cl_resp = (
        supabase.table("change_log")
        .select("*")
        .eq("event_id", event_id)
        .order("changed_at", desc=True)
        .limit(500)
        .execute()
    )
    changes: list[dict] = list(reversed(cl_resp.data or []))  # chronological order

    # --- Build workbook ---
    wb = Workbook()

    # Sheet 1: Master List (detailed)
    ws1 = wb.active
    _build_master_list_sheet(ws1, master_rows, conflict_ids, custom_fields)

    # Sheet 2: Data-quality analysis
    ws_q = wb.create_sheet()
    _build_quality_sheet(ws_q, analysis)

    # Sheet 3: Advice / recommendations
    ws_a = wb.create_sheet()
    _build_advice_sheet(ws_a, analysis)

    # Sheet 4: Exceptions
    ws2 = wb.create_sheet()
    _build_exceptions_sheet(ws2, exceptions)

    # Sheet 5: Summary
    ws3 = wb.create_sheet()
    _build_summary_sheet(ws3, run, user_id)

    # Sheet 6: Change Log
    ws4 = wb.create_sheet()
    _build_change_log_sheet(ws4, changes)

    # Serialise to bytes
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
